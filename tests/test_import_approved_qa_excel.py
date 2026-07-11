from __future__ import annotations

import hashlib
import json
from pathlib import Path

from openpyxl import Workbook, load_workbook

from rag_core.approved_qa import validate_approved_qa_records
from scripts.approved_qa_review import read_jsonl, review_validation_errors
from tools.import_approved_qa_excel import (
    EXIT_CONFIG,
    EXIT_INPUT,
    EXIT_PASSED,
    EXIT_VALIDATION,
    deterministic_qa_id,
    import_excel,
    main,
    _unsafe_text_issues,
)


JP_HEADERS = ["質問", "正解回答", "出典文書", "出典ページ", "テナント", "有効", "言い換え", "必須キーワード"]


def _write_xlsx(path: Path, rows: list[list], *, headers: list[str] | None = None,
                title: str = "QA入力", extra_sheet: bool = False) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title
    sheet.append(headers or JP_HEADERS)
    for row in rows:
        sheet.append(row)
    if extra_sheet:
        other = workbook.create_sheet("別シート")
        other.append(headers or JP_HEADERS)
        for row in rows:
            other.append(row)
    workbook.save(path)


def _valid_row(question: str = "架空サービスの期限は？", answer: str = "期限は月末です。") -> list:
    return [question, answer, "manual.pdf", "1,2", "default", "有効", "別の聞き方は？", "月末"]


def _run(tmp_path: Path, rows: list[list], **kwargs):
    source = tmp_path / kwargs.pop("filename", "任意 名称（QA）.xlsx")
    _write_xlsx(source, rows, **kwargs.pop("writer", {}))
    output = tmp_path / "out"
    summary = import_excel(input_path=source, output_dir=output, dry_run=True,
                           existing_approved_qa=None, **kwargs)
    return source, output, summary


def _jsonl(path: Path) -> list[dict]:
    return [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines() if line]


def test_normal_japanese_xlsx_arbitrary_unicode_and_spaces(tmp_path):
    _, output, summary = _run(tmp_path, [_valid_row()])
    candidates = _jsonl(output / "valid_candidates.jsonl")
    assert summary["status"] == "passed"
    assert len(candidates) == 1
    assert candidates[0]["status"] == "draft"
    assert candidates[0]["normalized_question"] == "架空サービスの期限は?"
    assert candidates[0]["approved_citations"][0]["source_pages"] == [1, 2]
    assert "approved_aliases" not in candidates[0]
    assert candidates[0]["candidate_metadata"]["aliases"] == ["別の聞き方は？"]
    assert summary["alias_candidate_count"] == 1
    assert summary["alias_conflict_count"] == 0
    assert validate_approved_qa_records(candidates) == []
    assert review_validation_errors(read_jsonl(output / "valid_candidates.jsonl")) == []


def test_sheet_name_and_zero_based_sheet_index(tmp_path):
    source = tmp_path / "sheets.xlsx"
    _write_xlsx(source, [_valid_row()], title="対象", extra_sheet=True)
    one = import_excel(input_path=source, output_dir=tmp_path / "by_name", sheet_name="対象",
                       dry_run=True, existing_approved_qa=None)
    two = import_excel(input_path=source, output_dir=tmp_path / "by_index", sheet_index=1,
                       dry_run=True, existing_approved_qa=None)
    assert one["valid_candidates"] == two["valid_candidates"] == 1


def test_mapping_json_cli_priority_and_custom_headers(tmp_path):
    headers = ["問い合わせ", "返答", "資料", "ページ番号", "顧客"]
    source = tmp_path / "custom.xlsx"
    _write_xlsx(source, [["Q", "A", "doc.pdf", "3", "tenant_a"]], headers=headers)
    mapping = tmp_path / "mapping.json"
    mapping.write_text(json.dumps({"columns": {"問い合わせ": "notes", "返答": "approved_answer", "資料": "source_doc",
                                                        "ページ番号": "source_pages", "顧客": "tenant_id"}}, ensure_ascii=False), encoding="utf-8")
    summary = import_excel(input_path=source, output_dir=tmp_path / "out", mapping_file=mapping,
                           cli_maps=["問い合わせ=question"], dry_run=True, existing_approved_qa=None)
    resolved = json.loads((tmp_path / "out/resolved_column_mapping.json").read_text(encoding="utf-8"))
    assert summary["valid_candidates"] == 1
    assert resolved["columns"]["問い合わせ"] == "question"


def test_mapping_conflicts_and_missing_required_column_are_validation_errors(tmp_path):
    source = tmp_path / "badmap.xlsx"
    _write_xlsx(source, [["Q", "A", "doc.pdf", "1"]], headers=["質問", "問い", "出典文書", "出典ページ"])
    output = tmp_path / "out"
    summary = import_excel(input_path=source, output_dir=output, dry_run=True, existing_approved_qa=None)
    assert summary["status"] == "failed"
    assert summary["error_count"] >= 1


def test_row_validation_empty_question_answer_invalid_page_boolean_and_path(tmp_path):
    rows = [
        ["", "A", "doc.pdf", "1", "default", "有効", "", ""],
        ["Q2", "", "doc.pdf", "1", "default", "有効", "", ""],
        ["Q3", "A", "doc.pdf", "0", "default", "有効", "", ""],
        ["Q4", "A", "doc.pdf", "1", "default", "maybe", "", ""],
        ["Q5", "A", "../secret.pdf", "1", "default", "有効", "", ""],
        ["Q6", "A", "/etc/passwd", "1", "default", "有効", "", ""],
    ]
    _, output, summary = _run(tmp_path, rows)
    assert summary["valid_candidates"] == 0
    assert summary["invalid_rows"] == 6
    codes = {error["code"] for row in _jsonl(output / "invalid_rows.jsonl") for error in row["errors"]}
    assert "row_validation" in codes


def test_malformed_list_and_json_are_invalid(tmp_path):
    _, output, summary = _run(tmp_path, [["Q", "A", "doc.pdf", "1", "default", "有効", '["broken"', "[]"]])
    assert summary["status"] == "failed"
    assert "malformed JSON" in (output / "invalid_rows.jsonl").read_text(encoding="utf-8")


def test_formula_and_control_character_detection(tmp_path):
    source = tmp_path / "formula.xlsx"
    _write_xlsx(source, [["=1+1", "A", "doc.pdf", "1", "default", "有効", "", ""]])
    output = tmp_path / "out"
    summary = import_excel(input_path=source, output_dir=output, dry_run=True, existing_approved_qa=None)
    assert summary["status"] == "failed"
    assert "formula_cell" in (output / "invalid_rows.jsonl").read_text(encoding="utf-8")
    assert {item["code"] for item in _unsafe_text_issues("bad\x01text", row=2, field="question")} == {"control_character"}
    assert "cell_too_large" in {item["code"] for item in _unsafe_text_issues("x" * 20_001, row=2, field="question")}


def test_invalid_status_and_external_link_are_rejected(tmp_path):
    headers = JP_HEADERS + ["状態"]
    source = tmp_path / "unsafe.xlsx"
    _write_xlsx(source, [[*_valid_row(), "approved"]], headers=headers)
    workbook = load_workbook(source)
    workbook["QA入力"]["A2"].hyperlink = "https://example.invalid/qa"
    workbook.save(source)
    output = tmp_path / "out"
    summary = import_excel(input_path=source, output_dir=output, dry_run=True, existing_approved_qa=None)
    text = (output / "invalid_rows.jsonl").read_text(encoding="utf-8")
    assert summary["status"] == "failed"
    assert "external_link" in text
    plain = tmp_path / "invalid_status.xlsx"
    _write_xlsx(plain, [[*_valid_row(), "approved"]], headers=headers)
    import_excel(input_path=plain, output_dir=tmp_path / "status_out", dry_run=True, existing_approved_qa=None)
    assert "candidate-only" in (tmp_path / "status_out/invalid_rows.jsonl").read_text(encoding="utf-8")


def test_duplicate_id_question_normalized_answer_conflict_and_alias_collision(tmp_path):
    headers = JP_HEADERS + ["QA ID"]
    rows = [
        ["同じ質問？", "A", "a.pdf", "1", "default", "有効", "別質問？", "", "same"],
        ["同じ質問?", "B", "a.pdf", "1", "default", "有効", "x", "", "same"],
        ["別質問？", "C", "a.pdf", "1", "default", "有効", "x,x", "", "other"],
    ]
    _, output, summary = _run(tmp_path, rows, writer={"headers": headers})
    text = (output / "invalid_rows.jsonl").read_text(encoding="utf-8")
    assert summary["status"] == "failed"
    assert "duplicate_qa_id" in text
    assert "answer_conflict" in text
    assert "alias_question_collision" in text or "duplicate_alias" in text


def test_existing_approved_qa_id_question_and_alias_collisions(tmp_path):
    existing = tmp_path / "approved.jsonl"
    existing.write_text(json.dumps({"qa_id": "existing", "question": "既存質問？", "normalized_question": "既存質問?",
                                    "approved_answer": "既存回答", "approved_citations": [{"source_doc": "p.pdf", "source_pages": [1]}],
                                    "tenant_id": "default", "language": "ja", "status": "approved"}, ensure_ascii=False) + "\n", encoding="utf-8")
    source = tmp_path / "input.xlsx"
    _write_xlsx(source, [["新規質問？", "A", "doc.pdf", "1", "default", "有効", "既存質問？", ""]], headers=JP_HEADERS + [])
    output = tmp_path / "out"
    summary = import_excel(input_path=source, output_dir=output, dry_run=True, existing_approved_qa=existing)
    assert summary["status"] == "failed"
    assert "alias_existing_collision" in (output / "invalid_rows.jsonl").read_text(encoding="utf-8")

    source2 = tmp_path / "id.xlsx"
    _write_xlsx(source2, [["既存質問？", "違う回答", "doc.pdf", "1", "default", "有効", "", "", "existing"]], headers=JP_HEADERS + ["QA ID"])
    summary2 = import_excel(input_path=source2, output_dir=tmp_path / "out2", dry_run=True, existing_approved_qa=existing)
    text = (tmp_path / "out2/invalid_rows.jsonl").read_text(encoding="utf-8")
    assert summary2["status"] == "failed"
    assert "existing_qa_id_conflict" in text and "existing_answer_conflict" in text


def test_existing_formal_alias_conflicts_with_xlsx_candidate_alias(tmp_path):
    existing = tmp_path / "approved_alias.jsonl"
    row = {"qa_id": "existing", "question": "正規質問？", "normalized_question": "正規質問?",
           "approved_answer": "回答", "approved_citations": [{"source_doc": "p.pdf", "source_pages": [1]}],
           "approved_aliases": ["登録済みalias？"], "tenant_id": "default", "language": "ja", "status": "approved"}
    existing.write_text(json.dumps(row, ensure_ascii=False) + "\n", encoding="utf-8")
    source = tmp_path / "candidate.xlsx"
    _write_xlsx(source, [["新規質問？", "新規回答", "doc.pdf", "1", "default", "有効", "登録済みalias?", ""]])
    output = tmp_path / "out_alias"
    summary = import_excel(input_path=source, output_dir=output, dry_run=True, existing_approved_qa=existing)
    assert summary["status"] == "failed"
    assert summary["alias_conflict_count"] == 1
    assert "alias_existing_collision" in (output / "invalid_rows.jsonl").read_text(encoding="utf-8")


def test_deterministic_id_same_input_and_meaningful_changes():
    base = deterministic_qa_id("t", "Q?", '{"source_doc":"a","source_pages":[1]}', "A")
    assert base == deterministic_qa_id("t", "Q?", '{"source_doc":"a","source_pages":[1]}', "A")
    assert base != deterministic_qa_id("t", "Q?", '{"source_doc":"a","source_pages":[2]}', "A")
    assert base != deterministic_qa_id("t", "Q?", '{"source_doc":"a","source_pages":[1]}', "B")


def test_corpus_warnings_and_strict_exit(tmp_path):
    source = tmp_path / "input.xlsx"
    _write_xlsx(source, [_valid_row()])
    passed = main(["--input", str(source), "--output-dir", str(tmp_path / "normal"), "--dry-run",
                   "--existing-approved-qa", str(tmp_path / "missing.jsonl")])
    strict = main(["--input", str(source), "--output-dir", str(tmp_path / "strict"), "--dry-run", "--strict",
                   "--existing-approved-qa", str(tmp_path / "missing.jsonl")])
    assert passed == EXIT_PASSED
    assert strict == EXIT_VALIDATION


def test_valid_corpus_checks_pages_and_required_terms(tmp_path):
    corpus = tmp_path / "corpus.jsonl"
    corpus.write_text(json.dumps({"source_doc": "manual.pdf", "source_pages": [1, 2], "text": "期限は月末です"}, ensure_ascii=False) + "\n", encoding="utf-8")
    _, output, summary = _run(tmp_path, [_valid_row()], corpus_jsonl=corpus)
    assert summary["warning_count"] == 0
    assert (output / "warnings.jsonl").read_text(encoding="utf-8") == ""


def test_dry_run_does_not_modify_production_or_vectorstore(tmp_path):
    production = tmp_path / "production.jsonl"
    vector = tmp_path / "vectorstore.bin"
    production.write_bytes(b"production-sentinel\n")
    vector.write_bytes(b"vector-sentinel\n")
    before = {p: hashlib.sha256(p.read_bytes()).hexdigest() for p in (production, vector)}
    source = tmp_path / "input.xlsx"
    _write_xlsx(source, [_valid_row()])
    import_excel(input_path=source, output_dir=tmp_path / "out", dry_run=True, existing_approved_qa=None)
    assert before == {p: hashlib.sha256(p.read_bytes()).hexdigest() for p in (production, vector)}


def test_limits_unsupported_input_and_exit_codes(tmp_path):
    wrong = tmp_path / "input.xls"
    wrong.write_text("x", encoding="utf-8")
    assert main(["--input", str(wrong), "--output-dir", str(tmp_path / "o")]) == EXIT_INPUT

    source = tmp_path / "input.xlsx"
    _write_xlsx(source, [_valid_row()])
    assert main(["--input", str(source), "--output-dir", str(tmp_path / "size"),
                 "--max-file-size-mb", "0.000001"]) == EXIT_INPUT
    assert main(["--input", str(source), "--output-dir", str(tmp_path / "config"),
                 "--sheet-name", "missing"]) == EXIT_CONFIG


def test_all_output_files_exist_and_jsonl_is_parseable(tmp_path):
    _, output, _ = _run(tmp_path, [_valid_row()])
    expected = {"validation_summary.json", "valid_candidates.jsonl", "invalid_rows.jsonl", "warnings.jsonl",
                "resolved_column_mapping.json", "import_report.md", "input_manifest.json"}
    assert expected == {path.name for path in output.iterdir()}
    for filename in ("valid_candidates.jsonl", "invalid_rows.jsonl", "warnings.jsonl"):
        _jsonl(output / filename)


def test_template_is_regenerable_and_contains_required_sheets(tmp_path):
    from scripts.build_approved_qa_import_template import build_template

    output = tmp_path / "template.xlsx"
    build_template(output)
    workbook = load_workbook(output, read_only=False)
    assert workbook.sheetnames == ["QA入力", "入力説明", "値候補"]
    assert [cell.value for cell in workbook["QA入力"][1]][:4] == ["質問", "正解回答", "出典文書", "出典ページ"]
    assert workbook["QA入力"]["A2"].value.startswith("架空サービス")
    assert workbook["QA入力"].data_validations.count >= 2
