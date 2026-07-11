from __future__ import annotations

import argparse
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter, quote_sheetname
from openpyxl.worksheet.datavalidation import DataValidation


HEADERS = [
    "質問", "正解回答", "出典文書", "出典ページ", "出典タイトル", "分類",
    "テナント", "文書版", "有効", "言い換え", "除外質問", "必須キーワード", "備考",
]

EXAMPLE = [
    "架空サービスの申請期限はいつですか？",
    "架空サービスの申請期限は毎月末日です。",
    "sample_manual.pdf", "2", "架空サービス利用案内", "手続き", "default", "v1",
    "有効", "申請はいつまでですか？", "解約期限はいつですか？", "毎月末日",
    "この行は架空の入力例です。実データではありません。",
]

EXPLANATIONS = [
    ("質問", "利用者が入力する承認対象の質問", "必須", "架空サービスの申請期限は？", "数式、秘密情報、巨大な文章", "空欄、既存質問との重複"),
    ("正解回答", "承認候補となる回答本文", "必須", "申請期限は毎月末日です。", "数式、未確認の推測", "空欄、根拠との不整合"),
    ("出典文書", "根拠文書の識別名（相対名）", "必須", "sample_manual.pdf", "絶対パス、../、URL", "文書が見つからない"),
    ("出典ページ", "1始まりのページ番号。複数はカンマ区切り又はJSON配列", "必須", "2,3", "0、負数、文字列", "ページが存在しない"),
    ("出典タイトル", "引用表示用のタイトル", "任意", "架空サービス利用案内", "スクリプト/HTML", "タイトルの取り違え"),
    ("分類", "候補の業務カテゴリ。candidate tagへ変換", "任意", "手続き", "機密分類名", "候補値との表記揺れ"),
    ("テナント", "既存tenant contractに従う識別子", "必須", "default", "空白、パス、空欄", "別テナントの指定"),
    ("文書版", "根拠文書の版", "任意", "v1", "秘密情報", "版の記入漏れ"),
    ("有効", "候補metadataの有効フラグ", "任意", "有効", "未定義値", "有効/無効の誤記"),
    ("言い換え", "明示的レビュー対象の別表現。複数はカンマ区切り又はJSON配列", "任意", "申請はいつまで？", "自動生成した未確認表現", "別質問との衝突"),
    ("除外質問", "この回答を使わない質問例。candidate metadataのみ", "任意", "解約期限は？", "秘密情報", "言い換えとの混同"),
    ("必須キーワード", "根拠内で確認したい語。candidate metadataのみ", "任意", "毎月末日", "未確認の語", "根拠に存在しない"),
    ("備考", "レビュー担当者向けメモ", "任意", "法務確認済み", "パスワード等", "個人情報の記載"),
]


def build_template(output: Path) -> None:
    workbook = Workbook()
    workbook.properties.creator = "approved_qa_import_template_builder"
    workbook.properties.created = datetime(2026, 1, 1, tzinfo=timezone.utc)
    workbook.properties.modified = datetime(2026, 1, 1, tzinfo=timezone.utc)
    qa = workbook.active
    qa.title = "QA入力"
    qa.append(HEADERS)
    qa.append(EXAMPLE)
    qa.freeze_panes = "A2"
    qa.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}2"
    for cell in qa[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in qa[2]:
        cell.fill = PatternFill("solid", fgColor="FFF2CC")
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    widths = [36, 48, 24, 14, 28, 14, 16, 12, 10, 30, 30, 26, 38]
    for index, width in enumerate(widths, start=1):
        qa.column_dimensions[get_column_letter(index)].width = width
    qa.row_dimensions[2].height = 58

    help_sheet = workbook.create_sheet("入力説明")
    help_sheet.append(["列名", "意味", "必須/任意", "入力例", "禁止事項", "よくあるエラー"])
    for row in EXPLANATIONS:
        help_sheet.append(row)
    help_sheet.freeze_panes = "A2"
    help_sheet.auto_filter.ref = f"A1:F{help_sheet.max_row}"
    for cell in help_sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for column, width in enumerate([18, 42, 12, 34, 34, 34], start=1):
        help_sheet.column_dimensions[get_column_letter(column)].width = width
    for row in help_sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    values = workbook.create_sheet("値候補")
    columns = {
        "有効値": ["有効", "無効", "true", "false"],
        "category候補": ["手続き", "料金", "仕様", "契約", "サポート", "その他"],
        "language候補": ["ja", "en"],
        "status候補": ["draft"],
        "expected answer mode候補": ["approved_exact_match", "grounded", "abstain"],
    }
    for col, (header, items) in enumerate(columns.items(), start=1):
        values.cell(1, col, header)
        for row, item in enumerate(items, start=2):
            values.cell(row, col, item)
        values.column_dimensions[get_column_letter(col)].width = 30
    for cell in values[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")

    enabled_col = HEADERS.index("有効") + 1
    category_col = HEADERS.index("分類") + 1
    enabled_validation = DataValidation(
        type="list", formula1=f"{quote_sheetname(values.title)}!$A$2:$A$5", allow_blank=True
    )
    category_validation = DataValidation(
        type="list", formula1=f"{quote_sheetname(values.title)}!$B$2:$B$7", allow_blank=True
    )
    qa.add_data_validation(enabled_validation)
    qa.add_data_validation(category_validation)
    enabled_validation.add(f"{get_column_letter(enabled_col)}2:{get_column_letter(enabled_col)}10001")
    category_validation.add(f"{get_column_letter(category_col)}2:{get_column_letter(category_col)}10001")

    output.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output)


def main(argv: list[str] | None = None) -> int:
    parser = argparse.ArgumentParser(description="Build the non-engineer approved-QA XLSX import template.")
    parser.add_argument("--output", type=Path, default=Path("templates/approved_qa_import_template.xlsx"))
    args = parser.parse_args(argv)
    build_template(args.output)
    print(args.output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
