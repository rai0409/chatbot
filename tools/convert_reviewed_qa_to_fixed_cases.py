#!/usr/bin/env python3
import argparse
import ast
import json
import re
import unicodedata
from collections import Counter
from pathlib import Path
from typing import Any, Dict, List

from openpyxl import load_workbook


JP = r"ぁ-んァ-ヴー一-龥々〆〤"


def normalize_unicode(value: Any) -> str:
    if value is None:
        return ""
    s = str(value)
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    return unicodedata.normalize("NFKC", s)


def normalize_display_text(value: Any) -> str:
    s = normalize_unicode(value)
    s = s.replace("\u3000", " ")
    s = re.sub(r"\s*\n\s*", " ", s)
    s = re.sub(r"[ \t]+", " ", s)

    # PDF折り返し由来の日本語内スペースを削除
    s = re.sub(rf"([{JP}])\s+([{JP}])", r"\1\2", s)
    s = re.sub(rf"([{JP}])\s+([。、，．！？!?）」』])", r"\1\2", s)
    s = re.sub(rf"([「『（(])\s+([{JP}])", r"\1\2", s)

    s = re.sub(r"\s+", " ", s)
    return s.strip()


def normalize_search_text(value: Any) -> str:
    s = normalize_display_text(value)
    s = s.lower()
    s = re.sub(r"\s+", " ", s)
    return s.strip()


def normalize_eval_text(value: Any) -> str:
    # QA評価用。日本語/英語問わず空白差を無視する。
    s = normalize_search_text(value)
    s = re.sub(r"\s+", "", s)
    return s.strip()


def parse_list_cell(value: Any) -> List[str]:
    if value is None:
        return []

    if isinstance(value, list):
        return [normalize_display_text(x) for x in value if normalize_display_text(x)]

    s = str(value).strip()
    if not s:
        return []

    for loader in (json.loads, ast.literal_eval):
        try:
            obj = loader(s)
            if isinstance(obj, list):
                return [normalize_display_text(x) for x in obj if normalize_display_text(x)]
        except Exception:
            pass

    parts = [x.strip() for x in s.replace("、", ",").split(",")]
    return [normalize_display_text(x) for x in parts if normalize_display_text(x)]


def to_int(value: Any, default: int = 0) -> int:
    if value is None or value == "":
        return default
    try:
        return int(value)
    except Exception:
        try:
            return int(float(str(value)))
        except Exception:
            return default


def read_review_xlsx(path: Path, include_statuses: set[str]) -> List[Dict[str, Any]]:
    wb = load_workbook(path)
    ws = wb.active

    headers = [str(c.value or "").strip() for c in ws[1]]
    required = {"review_status", "question", "expected_answer"}
    missing = required - set(headers)
    if missing:
        raise SystemExit(f"{path}: missing columns: {sorted(missing)}")

    out: List[Dict[str, Any]] = []

    for excel_row_no, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        src = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
        status = normalize_display_text(src.get("review_status")).lower()

        if status not in include_statuses:
            continue

        question = normalize_display_text(src.get("question"))
        expected_answer = normalize_display_text(src.get("expected_answer"))

        if not question or not expected_answer:
            continue

        case_id = normalize_display_text(src.get("case_id")) or f"{path.stem}_{excel_row_no}"
        expected_all = parse_list_cell(src.get("expected_all"))
        expected_any = parse_list_cell(src.get("expected_any"))
        forbidden_any = parse_list_cell(src.get("forbidden_any"))

        source_quote = normalize_display_text(src.get("source_quote"))
        if not source_quote:
            source_quote = f"Q: {question}\nA: {expected_answer}"

        row_out = {
            "case_id": case_id,
            "review_status": status,
            "source_review_xlsx": str(path),
            "source_pdf": normalize_display_text(src.get("source_pdf")),
            "source_page": to_int(src.get("source_page"), 0),
            "source_no": normalize_display_text(src.get("source_no")),
            "source_topic": normalize_display_text(src.get("source_topic")),
            "question": question,
            "expected_answer": expected_answer,
            "source_quote": source_quote,
            "question_type": normalize_display_text(src.get("question_type")) or "qa_fact",
            "expected_all": expected_all,
            "expected_any": expected_any,
            "expected_min_hits": to_int(src.get("expected_min_hits"), 1 if expected_any else 0),
            "forbidden_any": forbidden_any,
            "confidence": normalize_display_text(src.get("confidence")) or "unknown",
            "notes": normalize_display_text(src.get("notes")),

            # 検索・評価用
            "question_search_text": normalize_search_text(question),
            "question_eval_text": normalize_eval_text(question),
            "expected_answer_search_text": normalize_search_text(expected_answer),
            "expected_answer_eval_text": normalize_eval_text(expected_answer),
            "expected_all_eval": [normalize_eval_text(x) for x in expected_all],
            "expected_any_eval": [normalize_eval_text(x) for x in expected_any],
            "forbidden_any_eval": [normalize_eval_text(x) for x in forbidden_any],
        }

        out.append(row_out)

    return out


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input-xlsx", nargs="+", required=True)
    parser.add_argument("--output-jsonl", default="artifacts/fixed_qa_eval/fixed_qa_cases.jsonl")
    parser.add_argument("--include-status", default="reviewed,edited")
    args = parser.parse_args()

    include_statuses = {x.strip().lower() for x in args.include_status.split(",") if x.strip()}
    output_jsonl = Path(args.output_jsonl)
    output_jsonl.parent.mkdir(parents=True, exist_ok=True)

    all_rows: List[Dict[str, Any]] = []
    seen = set()

    for input_path in args.input_xlsx:
        p = Path(input_path)
        if not p.exists():
            raise SystemExit(f"missing input xlsx: {p}")

        rows = read_review_xlsx(p, include_statuses)
        for r in rows:
            key = (r["question_eval_text"], r["expected_answer_eval_text"])
            if key in seen:
                continue
            seen.add(key)
            all_rows.append(r)

    with output_jsonl.open("w", encoding="utf-8") as f:
        for r in all_rows:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")

    by_pdf = Counter(r["source_pdf"] for r in all_rows)
    by_type = Counter(r["question_type"] for r in all_rows)
    by_status = Counter(r["review_status"] for r in all_rows)

    report_path = output_jsonl.parent / "fixed_qa_cases_report.md"
    lines = [
        "# Fixed QA Cases Report",
        "",
        f"- output_jsonl: {output_jsonl}",
        f"- total_cases: {len(all_rows)}",
        f"- include_statuses: {sorted(include_statuses)}",
        "",
        "## By PDF",
        "",
    ]

    for k, v in by_pdf.items():
        lines.append(f"- {k}: {v}")

    lines += ["", "## By Question Type", ""]
    for k, v in by_type.items():
        lines.append(f"- {k}: {v}")

    lines += ["", "## By Review Status", ""]
    for k, v in by_status.items():
        lines.append(f"- {k}: {v}")

    report_path.write_text("\n".join(lines), encoding="utf-8")

    print(json.dumps({
        "status": "ok" if all_rows else "no_cases",
        "total_cases": len(all_rows),
        "output_jsonl": str(output_jsonl),
        "report": str(report_path),
        "by_pdf": dict(by_pdf),
        "by_type": dict(by_type),
        "by_status": dict(by_status),
    }, ensure_ascii=False, indent=2))

    return 0 if all_rows else 2


if __name__ == "__main__":
    raise SystemExit(main())
