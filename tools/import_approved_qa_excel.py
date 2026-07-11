from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
import traceback
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path, PurePath
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from rag_core.approved_qa import validate_approved_qa_records
from rag_core.question_normalization import normalize_question_for_exact_match


EXIT_PASSED = 0
EXIT_VALIDATION = 1
EXIT_CONFIG = 2
EXIT_INPUT = 3
EXIT_INTERNAL = 4

DEFAULT_MAX_FILE_SIZE_MB = 20.0
DEFAULT_MAX_ROWS = 10_000
MAX_WORKSHEETS = 20
MAX_COLUMNS = 80
MAX_CELL_LENGTH = 20_000
MAX_ALIASES = 20
MAX_TERMS = 30
ALLOWED_STATUSES = {"draft", "approved", "rejected"}
ALLOWED_LANGUAGES = {"ja", "en"}
CANONICAL_FIELDS = {
    "qa_id", "question", "approved_answer", "approved_citations", "source_doc",
    "source_pages", "title", "chunk_id", "category", "tenant_id", "language",
    "doc_version", "status", "enabled", "aliases", "excluded_questions",
    "required_terms", "notes",
}
REQUIRED_FIELDS = {"question", "approved_answer", "source_doc"}

JAPANESE_ALIASES = {
    "QA ID": "qa_id", "QAID": "qa_id", "質問": "question", "問い": "question",
    "QA質問": "question", "正解": "approved_answer", "回答": "approved_answer",
    "正解回答": "approved_answer", "出典": "source_doc", "出典文書": "source_doc",
    "文書名": "source_doc", "ページ": "source_pages", "頁": "source_pages",
    "出典ページ": "source_pages", "出典タイトル": "title", "タイトル": "title",
    "チャンクID": "chunk_id", "分類": "category", "カテゴリ": "category",
    "テナント": "tenant_id", "言語": "language", "文書版": "doc_version",
    "状態": "status", "ステータス": "status", "有効": "enabled", "使用": "enabled",
    "言い換え": "aliases", "別表現": "aliases", "除外質問": "excluded_questions",
    "不正解質問": "excluded_questions", "必須語": "required_terms",
    "必須キーワード": "required_terms", "備考": "notes",
}

BOOL_TRUE = {"true", "1", "yes", "y", "on", "有効", "使用", "はい", "○"}
BOOL_FALSE = {"false", "0", "no", "n", "off", "無効", "不使用", "いいえ", "×"}
CONTROL_RE = re.compile(r"[\x00-\x08\x0b\x0c\x0e-\x1f\x7f]")
HTML_SCRIPT_RE = re.compile(r"<\s*(?:script|iframe|object|embed|svg)|javascript\s*:", re.I)
SECRET_RE = re.compile(
    r"(?:-----BEGIN [A-Z ]*PRIVATE KEY-----|sk-[A-Za-z0-9_-]{20,}|AKIA[0-9A-Z]{16}|"
    r"(?:api[_-]?key|client[_-]?secret|password)\s*[:=]\s*[^\s,;]{8,})",
    re.I,
)


class ConfigError(ValueError):
    pass


class InputError(ValueError):
    pass


def _now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def _text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _issue(severity: str, code: str, message: str, *, row: int | None = None,
           field: str | None = None, value: Any = None) -> dict:
    out = {"severity": severity, "code": code, "message": message}
    if row is not None:
        out["row"] = row
    if field:
        out["field"] = field
    if value not in (None, ""):
        out["value_preview"] = _text(value)[:160]
    return out


def parse_map_arg(value: str) -> tuple[str, str]:
    if "=" not in value:
        raise ConfigError(f"--map must be INPUT_COLUMN=canonical_field: {value}")
    source, target = (_text(part) for part in value.split("=", 1))
    if not source or target not in CANONICAL_FIELDS:
        raise ConfigError(f"invalid --map: {value}")
    return source, target


def load_mapping_file(path: Path | None) -> dict[str, str]:
    if path is None:
        return {}
    try:
        payload = json.loads(path.read_text(encoding="utf-8"))
    except Exception as exc:
        raise ConfigError(f"mapping file is not readable JSON: {exc}") from exc
    mapping = payload.get("columns") if isinstance(payload, dict) and "columns" in payload else payload
    if not isinstance(mapping, dict):
        raise ConfigError("mapping JSON must be an object or contain an object named 'columns'")
    out: dict[str, str] = {}
    for source, target in mapping.items():
        source_text, target_text = _text(source), _text(target)
        if not source_text or target_text not in CANONICAL_FIELDS:
            raise ConfigError(f"invalid mapping entry: {source!r} -> {target!r}")
        out[source_text] = target_text
    return out


def resolve_column_mapping(headers: list[str], *, file_mapping: dict[str, str],
                           cli_mapping: dict[str, str]) -> tuple[dict[str, str], list[dict]]:
    issues: list[dict] = []
    if any(not header for header in headers):
        issues.append(_issue("error", "blank_header", "Blank column headers are not allowed."))
    duplicates = [name for name, count in Counter(headers).items() if name and count > 1]
    if duplicates:
        issues.append(_issue("error", "duplicate_header", f"Duplicate headers: {duplicates}"))
    unknown_explicit = (set(file_mapping) | set(cli_mapping)) - set(headers)
    if unknown_explicit:
        issues.append(_issue("error", "mapping_column_missing", f"Mapped columns not found: {sorted(unknown_explicit)}"))

    resolved: dict[str, str] = {}
    for header in headers:
        target = None
        if header in cli_mapping:
            target = cli_mapping[header]
        elif header in file_mapping:
            target = file_mapping[header]
        elif header in CANONICAL_FIELDS:
            target = header
        elif header in JAPANESE_ALIASES:
            target = JAPANESE_ALIASES[header]
        if target:
            resolved[header] = target

    reverse: dict[str, list[str]] = defaultdict(list)
    for header, target in resolved.items():
        reverse[target].append(header)
    conflicts = {target: cols for target, cols in reverse.items() if len(cols) > 1}
    if conflicts:
        issues.append(_issue("error", "canonical_mapping_conflict", f"Multiple columns map to one field: {conflicts}"))
    missing = sorted(REQUIRED_FIELDS - set(resolved.values()))
    if missing:
        issues.append(_issue("error", "missing_required_column", f"Missing required canonical columns: {missing}"))
    return resolved, issues


def parse_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    text = _text(value).lower()
    if text in BOOL_TRUE:
        return True
    if text in BOOL_FALSE:
        return False
    raise ValueError("enabled must be a supported boolean value")


def parse_list(value: Any, *, field: str, limit: int) -> list[str]:
    if value in (None, ""):
        return []
    if isinstance(value, (list, tuple)):
        items = list(value)
    else:
        raw = _text(value)
        if raw.startswith("[") or raw.startswith("{"):
            try:
                parsed = json.loads(raw)
            except json.JSONDecodeError as exc:
                raise ValueError(f"malformed JSON in {field}: {exc.msg}") from exc
            if not isinstance(parsed, list):
                raise ValueError(f"{field} JSON must be a list")
            items = parsed
        else:
            items = re.split(r"[\n,、]+", raw)
    result = [_text(item) for item in items if _text(item)]
    if len(result) > limit:
        raise ValueError(f"{field} exceeds limit {limit}")
    return result


def parse_pages(value: Any) -> list[int]:
    values = parse_list(value, field="source_pages", limit=100)
    pages: list[int] = []
    for item in values:
        if re.fullmatch(r"\d+", item) is None or int(item) < 1:
            raise ValueError(f"invalid source page: {item}")
        pages.append(int(item))
    if not pages:
        raise ValueError("source_pages is required and must contain positive integers")
    return pages


def safe_source_doc(value: Any) -> str:
    source_doc = _text(value)
    if not source_doc:
        raise ValueError("source_doc is required")
    normalized = source_doc.replace("\\", "/")
    if Path(source_doc).is_absolute() or re.match(r"^[A-Za-z]:[/\\]", source_doc):
        raise ValueError("absolute source_doc paths are not allowed")
    if ".." in PurePath(normalized).parts:
        raise ValueError("source_doc path traversal is not allowed")
    return source_doc


def deterministic_qa_id(tenant_id: str, normalized_question: str, source_identity: str,
                        approved_answer: str) -> str:
    answer_fingerprint = hashlib.sha256(approved_answer.encode("utf-8")).hexdigest()
    payload = "\n".join([tenant_id, normalized_question, source_identity, answer_fingerprint])
    return "qa_" + hashlib.sha256(payload.encode("utf-8")).hexdigest()[:16]


def _unsafe_text_issues(value: Any, *, row: int, field: str, formula: bool = False) -> list[dict]:
    text = _text(value)
    issues: list[dict] = []
    if formula or (text and text[0] in "=+-@" and not text.startswith("++")):
        issues.append(_issue("error", "formula_cell", "Formula or formula-like cell is not accepted.", row=row, field=field))
    if len(text) > MAX_CELL_LENGTH:
        issues.append(_issue("error", "cell_too_large", f"Cell exceeds {MAX_CELL_LENGTH} characters.", row=row, field=field))
    if "\x00" in text:
        issues.append(_issue("error", "null_byte", "Null bytes are not allowed.", row=row, field=field))
    elif CONTROL_RE.search(text) or ILLEGAL_CHARACTERS_RE.search(text):
        issues.append(_issue("error", "control_character", "Control characters are not allowed.", row=row, field=field))
    if HTML_SCRIPT_RE.search(text):
        issues.append(_issue("warning", "active_content_suspected", "Script or active HTML-like content detected.", row=row, field=field))
    if SECRET_RE.search(text):
        issues.append(_issue("warning", "secret_like_value", "Secret-like pattern detected; inspect before review.", row=row, field=field))
    return issues


def read_existing(path: Path | None) -> tuple[list[dict], list[dict]]:
    if path is None or not path.exists():
        warning = [] if path is None else [_issue("warning", "existing_qa_unavailable", f"Existing approved QA was not found: {path}")]
        return [], warning
    try:
        records = [json.loads(line) for line in path.read_text(encoding="utf-8").splitlines() if line.strip()]
    except Exception as exc:
        raise ConfigError(f"existing approved QA is unreadable: {exc}") from exc
    return records, []


def _existing_indexes(records: Iterable[dict]) -> tuple[dict[str, dict], dict[tuple[str, str], dict]]:
    ids: dict[str, dict] = {}
    questions: dict[tuple[str, str], dict] = {}
    for record in records:
        if not isinstance(record, dict):
            continue
        qa_id = _text(record.get("qa_id"))
        tenant = _text(record.get("tenant_id")) or "default"
        normalized = normalize_question_for_exact_match(_text(record.get("normalized_question") or record.get("question")))
        if qa_id:
            ids[qa_id] = record
        if normalized:
            questions[(tenant, normalized)] = record
        for alias in record.get("approved_aliases") or []:
            if isinstance(alias, str) and normalize_question_for_exact_match(alias):
                questions[(tenant, normalize_question_for_exact_match(alias))] = record
    return ids, questions


def _load_corpus(path: Path | None) -> dict[str, dict[int, str]]:
    if path is None:
        return {}
    result: dict[str, dict[int, str]] = defaultdict(dict)
    try:
        for line in path.read_text(encoding="utf-8").splitlines():
            if not line.strip():
                continue
            row = json.loads(line)
            doc = _text(row.get("source_doc"))
            pages = row.get("source_pages") or []
            text = _text(row.get("display_text") or row.get("text"))
            if doc:
                for page in pages if isinstance(pages, list) else []:
                    if isinstance(page, int):
                        result[doc][page] = result[doc].get(page, "") + " " + text
    except Exception as exc:
        raise ConfigError(f"corpus JSONL is unreadable: {exc}") from exc
    return result


def _candidate_from_row(values: dict[str, Any], *, row_number: int, formulas: set[str],
                        external_links: set[str],
                        timestamp: str) -> tuple[dict | None, list[dict], list[dict]]:
    errors: list[dict] = []
    warnings: list[dict] = []
    for field, value in values.items():
        for issue in _unsafe_text_issues(value, row=row_number, field=field, formula=field in formulas):
            (errors if issue["severity"] == "error" else warnings).append(issue)
    for field in sorted(external_links):
        errors.append(_issue("error", "external_link", "External hyperlinks are not accepted in QA intake.", row=row_number, field=field))
    if errors:
        return None, errors, warnings
    try:
        question = _text(values.get("question"))
        answer = _text(values.get("approved_answer"))
        if not question:
            raise ValueError("question is required")
        if not answer:
            raise ValueError("approved_answer is required")
        tenant = _text(values.get("tenant_id"))
        if not tenant or not re.fullmatch(r"[A-Za-z0-9][A-Za-z0-9._-]{0,127}", tenant):
            raise ValueError("tenant_id is required and must use letters, digits, dot, underscore, or hyphen")
        source_doc = safe_source_doc(values.get("source_doc"))
        pages = parse_pages(values.get("source_pages"))
        status = _text(values.get("status")) or "draft"
        if status not in ALLOWED_STATUSES:
            raise ValueError(f"invalid status: {status}")
        if status != "draft":
            raise ValueError("Excel intake is candidate-only; status must be draft")
        language = _text(values.get("language")) or "ja"
        if language not in ALLOWED_LANGUAGES:
            raise ValueError(f"unsupported language: {language}")
        enabled = parse_bool(values.get("enabled") if values.get("enabled") not in (None, "") else True)
        aliases = parse_list(values.get("aliases"), field="aliases", limit=MAX_ALIASES)
        excluded = parse_list(values.get("excluded_questions"), field="excluded_questions", limit=MAX_ALIASES)
        required_terms = parse_list(values.get("required_terms"), field="required_terms", limit=MAX_TERMS)
        category = _text(values.get("category"))
        normalized = normalize_question_for_exact_match(question)
        source_identity = json.dumps({"source_doc": source_doc, "source_pages": pages}, ensure_ascii=False, sort_keys=True)
        qa_id = _text(values.get("qa_id")) or deterministic_qa_id(tenant, normalized, source_identity, answer)
        citation = {"source_doc": source_doc, "source_pages": pages}
        if _text(values.get("title")):
            citation["title"] = _text(values.get("title"))
        if _text(values.get("chunk_id")):
            citation["chunk_id"] = _text(values.get("chunk_id"))
        record = {
            "qa_id": qa_id,
            "question": question,
            "normalized_question": normalized,
            "approved_answer": answer,
            "approved_citations": [citation],
            "tags": [tag for tag in [category, "excel_candidate"] if tag],
            "language": language,
            "tenant_id": tenant,
            "doc_version": _text(values.get("doc_version")),
            "status": "draft",
            "created_at": timestamp,
            "notes": _text(values.get("notes")),
            "candidate_metadata": {
                "source": "approved_qa_excel_import",
                "source_row": row_number,
                "enabled": enabled,
                "aliases": aliases,
                "excluded_questions": excluded,
                "required_terms": required_terms,
            },
        }
        return record, errors, warnings
    except Exception as exc:
        errors.append(_issue("error", "row_validation", str(exc), row=row_number))
        return None, errors, warnings


def _cross_validate(candidates: list[dict], existing: list[dict], corpus: dict[str, dict[int, str]]) -> tuple[set[int], list[dict], list[dict]]:
    bad: set[int] = set()
    errors: list[dict] = []
    warnings: list[dict] = []
    existing_ids, existing_questions = _existing_indexes(existing)
    seen_ids: dict[str, int] = {}
    seen_questions: dict[tuple[str, str], tuple[int, str]] = {}
    seen_aliases: dict[tuple[str, str], int] = {}
    primary_questions = {(c["tenant_id"], c["normalized_question"]): i for i, c in enumerate(candidates)}

    def err(i: int, code: str, message: str) -> None:
        bad.add(i)
        errors.append(_issue("error", code, message, row=candidates[i]["candidate_metadata"]["source_row"]))

    for i, record in enumerate(candidates):
        qa_id, tenant, normalized = record["qa_id"], record["tenant_id"], record["normalized_question"]
        key = (tenant, normalized)
        if qa_id in seen_ids:
            err(i, "duplicate_qa_id", f"Duplicate qa_id with row {candidates[seen_ids[qa_id]]['candidate_metadata']['source_row']}: {qa_id}")
            bad.add(seen_ids[qa_id])
        else:
            seen_ids[qa_id] = i
        if key in seen_questions:
            previous_i, previous_answer = seen_questions[key]
            code = "answer_conflict" if previous_answer != record["approved_answer"] else "normalized_duplicate"
            err(i, code, f"Duplicate normalized question with row {candidates[previous_i]['candidate_metadata']['source_row']}")
            bad.add(previous_i)
        else:
            seen_questions[key] = (i, record["approved_answer"])
        if qa_id in existing_ids:
            err(i, "existing_qa_id_conflict", f"qa_id already exists: {qa_id}")
        if key in existing_questions:
            existing_answer = _text(existing_questions[key].get("approved_answer"))
            code = "existing_answer_conflict" if existing_answer != record["approved_answer"] else "existing_question_duplicate"
            err(i, code, "Question conflicts with existing approved QA in the same tenant.")

        local_aliases: set[str] = set()
        for alias in record["candidate_metadata"]["aliases"]:
            alias_normalized = normalize_question_for_exact_match(alias)
            alias_key = (tenant, alias_normalized)
            if not alias_normalized:
                err(i, "empty_alias", "Alias becomes empty after normalization.")
            elif alias_normalized in local_aliases:
                err(i, "duplicate_alias", f"Alias is duplicated within the row: {alias}")
            elif alias_key == key or alias_key in primary_questions and primary_questions[alias_key] != i:
                err(i, "alias_question_collision", f"Alias collides with a candidate question: {alias}")
            elif alias_key in existing_questions:
                err(i, "alias_existing_collision", f"Alias collides with existing approved QA: {alias}")
            elif alias_key in seen_aliases and seen_aliases[alias_key] != i:
                err(i, "duplicate_alias", f"Alias duplicates another row: {alias}")
            else:
                seen_aliases[alias_key] = i
            local_aliases.add(alias_normalized)

        citation = record["approved_citations"][0]
        source_doc, pages = citation["source_doc"], citation["source_pages"]
        row_no = record["candidate_metadata"]["source_row"]
        if corpus:
            if source_doc not in corpus:
                warnings.append(_issue("warning", "source_doc_not_found", "source_doc not found in supplied corpus.", row=row_no, field="source_doc"))
            else:
                missing_pages = [page for page in pages if page not in corpus[source_doc]]
                if missing_pages:
                    warnings.append(_issue("warning", "source_pages_not_found", f"Pages not found in supplied corpus: {missing_pages}", row=row_no, field="source_pages"))
                corpus_text = " ".join(corpus[source_doc].get(page, "") for page in pages)
                missing_terms = [term for term in record["candidate_metadata"]["required_terms"] if term not in corpus_text]
                if missing_terms:
                    warnings.append(_issue("warning", "required_terms_not_found", f"Required terms not found: {missing_terms}", row=row_no))
        else:
            warnings.append(_issue("warning", "corpus_validation_not_run", "Source/page/required-term existence was not verifiable; no --corpus-jsonl was supplied.", row=row_no))
    return bad, errors, warnings


def _write_json(path: Path, payload: Any) -> None:
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def _write_jsonl(path: Path, rows: Iterable[dict]) -> None:
    with path.open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(row, ensure_ascii=False, separators=(",", ":")) + "\n")


def import_excel(*, input_path: Path, output_dir: Path, sheet_name: str | None = None,
                 sheet_index: int | None = None, mapping_file: Path | None = None,
                 cli_maps: Iterable[str] = (), existing_approved_qa: Path | None = None,
                 corpus_jsonl: Path | None = None, max_rows: int = DEFAULT_MAX_ROWS,
                 max_file_size_mb: float = DEFAULT_MAX_FILE_SIZE_MB, strict: bool = False,
                 dry_run: bool = False) -> dict:
    del dry_run  # All modes are intentionally candidate-only; retained as an explicit operator signal.
    if input_path.suffix.lower() != ".xlsx":
        raise InputError("Only .xlsx is supported; .xls, .xlsm, and .ods are not accepted.")
    if not input_path.is_file():
        raise InputError(f"Input file not found: {input_path}")
    if input_path.stat().st_size > int(max_file_size_mb * 1024 * 1024):
        raise InputError(f"Input exceeds --max-file-size-mb ({max_file_size_mb}).")
    if max_rows < 1 or max_file_size_mb <= 0:
        raise ConfigError("row and file-size limits must be positive")
    if sheet_name is not None and sheet_index is not None:
        raise ConfigError("Specify only one of --sheet-name or --sheet-index")

    try:
        workbook = load_workbook(input_path, read_only=False, data_only=False, keep_links=True)
    except Exception as exc:
        raise InputError(f"Unreadable XLSX: {exc}") from exc
    if len(workbook.worksheets) > MAX_WORKSHEETS:
        raise InputError(f"Workbook exceeds maximum worksheet count ({MAX_WORKSHEETS}).")
    if sheet_name is not None:
        if sheet_name not in workbook.sheetnames:
            raise ConfigError(f"Worksheet not found: {sheet_name}")
        sheet = workbook[sheet_name]
    elif sheet_index is not None:
        if sheet_index < 0 or sheet_index >= len(workbook.worksheets):
            raise ConfigError(f"--sheet-index is zero-based and out of range: {sheet_index}")
        sheet = workbook.worksheets[sheet_index]
    else:
        sheet = workbook["QA入力"] if "QA入力" in workbook.sheetnames else workbook.worksheets[0]
    if sheet.max_column > MAX_COLUMNS:
        raise InputError(f"Worksheet exceeds maximum columns ({MAX_COLUMNS}).")
    if max(0, sheet.max_row - 1) > max_rows:
        raise InputError(f"Worksheet exceeds --max-rows ({max_rows}).")

    headers = [_text(sheet.cell(1, col).value) for col in range(1, sheet.max_column + 1)]
    file_mapping = load_mapping_file(mapping_file)
    cli_mapping = dict(parse_map_arg(item) for item in cli_maps)
    resolved, mapping_issues = resolve_column_mapping(headers, file_mapping=file_mapping, cli_mapping=cli_mapping)
    errors = [issue for issue in mapping_issues if issue["severity"] == "error"]
    warnings = [issue for issue in mapping_issues if issue["severity"] == "warning"]

    timestamp = _now()
    candidates: list[dict] = []
    invalid_rows: list[dict] = []
    if not errors:
        for row_no in range(2, sheet.max_row + 1):
            row_cells = [sheet.cell(row_no, col) for col in range(1, sheet.max_column + 1)]
            if all(cell.value in (None, "") for cell in row_cells):
                continue
            values: dict[str, Any] = {}
            formulas: set[str] = set()
            external_links: set[str] = set()
            for col, header in enumerate(headers, start=1):
                if header not in resolved:
                    continue
                target = resolved[header]
                cell = sheet.cell(row_no, col)
                values[target] = cell.value
                if cell.data_type == "f":
                    formulas.add(target)
                if cell.hyperlink and _text(cell.hyperlink.target).lower().startswith(("http://", "https://", "file://")):
                    external_links.add(target)
            candidate, row_errors, row_warnings = _candidate_from_row(
                values, row_number=row_no, formulas=formulas, external_links=external_links, timestamp=timestamp
            )
            warnings.extend(row_warnings)
            if candidate is None:
                invalid_rows.append({"row": row_no, "errors": row_errors})
                errors.extend(row_errors)
            else:
                candidates.append(candidate)

    existing, existing_warnings = read_existing(existing_approved_qa)
    warnings.extend(existing_warnings)
    corpus = _load_corpus(corpus_jsonl)
    bad_indexes, cross_errors, cross_warnings = _cross_validate(candidates, existing, corpus)
    errors.extend(cross_errors)
    warnings.extend(cross_warnings)
    for i in sorted(bad_indexes):
        invalid_rows.append({"row": candidates[i]["candidate_metadata"]["source_row"], "candidate": candidates[i],
                             "errors": [issue for issue in cross_errors if issue.get("row") == candidates[i]["candidate_metadata"]["source_row"]]})
    valid = [record for i, record in enumerate(candidates) if i not in bad_indexes]

    validator_errors = validate_approved_qa_records(valid)
    if validator_errors:
        for message in validator_errors:
            errors.append(_issue("error", "existing_validator", message))
        valid = []

    output_dir.mkdir(parents=True, exist_ok=True)
    _write_jsonl(output_dir / "valid_candidates.jsonl", valid)
    _write_jsonl(output_dir / "invalid_rows.jsonl", invalid_rows)
    _write_jsonl(output_dir / "warnings.jsonl", warnings)
    _write_json(output_dir / "resolved_column_mapping.json", {"columns": resolved, "precedence": ["cli", "mapping_json", "canonical_exact", "japanese_alias"]})
    input_hash = hashlib.sha256(input_path.read_bytes()).hexdigest()
    manifest = {"input_file": input_path.name, "input_sha256": input_hash, "input_size_bytes": input_path.stat().st_size,
                "worksheet": sheet.title, "worksheet_index": workbook.worksheets.index(sheet), "processed_at": timestamp,
                "candidate_only": True, "production_modified": False, "limits": {"max_file_size_mb": max_file_size_mb,
                "max_worksheets": MAX_WORKSHEETS, "max_rows": max_rows, "max_columns": MAX_COLUMNS,
                "max_cell_length": MAX_CELL_LENGTH, "max_aliases": MAX_ALIASES, "max_terms": MAX_TERMS}}
    _write_json(output_dir / "input_manifest.json", manifest)
    alias_count = sum(len(record["candidate_metadata"]["aliases"]) for record in candidates)
    alias_conflict_count = sum(
        1 for issue in errors
        if issue.get("code") in {"alias_question_collision", "alias_existing_collision", "duplicate_alias", "empty_alias"}
    )
    summary = {"status": "failed" if errors or (strict and warnings) else "passed", "input_rows": len(candidates) + len([r for r in invalid_rows if "candidate" not in r]),
               "valid_candidates": len(valid), "invalid_rows": len({row["row"] for row in invalid_rows}),
               "error_count": len(errors), "warning_count": len(warnings), "strict": strict,
               "alias_candidate_count": alias_count, "alias_conflict_count": alias_conflict_count,
               "candidate_only": True, "production_modified": False}
    _write_json(output_dir / "validation_summary.json", summary)
    report = ["# Approved QA Excel Import Report", "", f"- Status: **{summary['status']}**",
              f"- Worksheet: `{sheet.title}`", f"- Valid candidates: {len(valid)}",
              f"- Invalid rows: {summary['invalid_rows']}", f"- Errors: {len(errors)}", f"- Warnings: {len(warnings)}",
              f"- Alias candidates: {alias_count}", f"- Alias conflicts: {alias_conflict_count}",
              "- Output is candidate-only; production approved QA and vector collections were not modified.", "",
              "## Next step", "", "Review `valid_candidates.jsonl` with `scripts/approved_qa_review.py`. Promotion/export is a separate explicit operation."]
    (output_dir / "import_report.md").write_text("\n".join(report) + "\n", encoding="utf-8")
    return summary


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Safely import approved-QA candidates from an XLSX workbook.")
    parser.add_argument("--input", required=True, type=Path, dest="input_path")
    parser.add_argument("--sheet-name")
    parser.add_argument("--sheet-index", type=int)
    parser.add_argument("--mapping-file", type=Path)
    parser.add_argument("--map", action="append", default=[], dest="cli_maps")
    parser.add_argument("--output-dir", required=True, type=Path)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--strict", action="store_true")
    parser.add_argument("--max-rows", type=int, default=DEFAULT_MAX_ROWS)
    parser.add_argument("--max-file-size-mb", type=float, default=DEFAULT_MAX_FILE_SIZE_MB)
    parser.add_argument("--existing-approved-qa", type=Path, default=Path("data/approved_qa/default.jsonl"))
    parser.add_argument("--corpus-jsonl", type=Path)
    return parser


def main(argv: list[str] | None = None) -> int:
    parser = build_parser()
    try:
        args = parser.parse_args(argv)
        summary = import_excel(**vars(args))
        print(json.dumps(summary, ensure_ascii=False, sort_keys=True))
        return EXIT_VALIDATION if summary["status"] == "failed" else EXIT_PASSED
    except ConfigError as exc:
        print(f"CONFIG ERROR: {exc}", file=sys.stderr)
        return EXIT_CONFIG
    except InputError as exc:
        print(f"INPUT ERROR: {exc}", file=sys.stderr)
        return EXIT_INPUT
    except SystemExit:
        raise
    except Exception as exc:
        print(f"INTERNAL ERROR: {exc}", file=sys.stderr)
        if "--debug" in (argv or []):
            traceback.print_exc()
        return EXIT_INTERNAL


if __name__ == "__main__":
    raise SystemExit(main())
