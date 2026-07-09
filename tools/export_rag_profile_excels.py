#!/usr/bin/env python3
"""Export editable RAG profile Excel files to JSON/JSONL runtime files."""

from __future__ import annotations

import argparse
import json
import math
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


DEFAULT_INPUT_DIR = Path("config/rag_profiles/volcano_demo/source")
DEFAULT_OUTPUT_DIR = Path("config/rag_profiles/volcano_demo/exported")


class ExportError(ValueError):
    """Raised when a workbook cannot be exported safely."""


@dataclass(frozen=True)
class ExportSpec:
    source_name: str
    output_name: str
    required_columns: tuple[str, ...]
    root_key: str | None
    list_columns: frozenset[str] = frozenset()
    int_columns: frozenset[str] = frozenset()
    bool_columns: frozenset[str] = frozenset({"enabled"})
    jsonl: bool = False


EXPORT_SPECS: tuple[ExportSpec, ...] = (
    ExportSpec(
        source_name="01_profile.xlsx",
        output_name="profile.json",
        required_columns=("key", "value", "description"),
        root_key=None,
    ),
    ExportSpec(
        source_name="02_question_types.xlsx",
        output_name="question_type_rules.json",
        required_columns=(
            "type_id",
            "description",
            "question_contains_any",
            "question_contains_all",
            "priority",
            "enabled",
        ),
        root_key="rules",
        list_columns=frozenset({"question_contains_any", "question_contains_all"}),
        int_columns=frozenset({"priority"}),
    ),
    ExportSpec(
        source_name="03_domain_terms.xlsx",
        output_name="domain_terms.json",
        required_columns=(
            "term_id",
            "category",
            "term",
            "aliases",
            "weight",
            "negative_weight",
            "description",
            "enabled",
        ),
        root_key="terms",
        list_columns=frozenset({"aliases"}),
        int_columns=frozenset({"weight", "negative_weight"}),
    ),
    ExportSpec(
        source_name="04_synonyms.xlsx",
        output_name="synonyms.json",
        required_columns=("canonical", "synonyms", "description", "enabled"),
        root_key="synonyms",
        list_columns=frozenset({"synonyms"}),
    ),
    ExportSpec(
        source_name="05_retrieval_boost_rules.xlsx",
        output_name="retrieval_boost_rules.json",
        required_columns=(
            "rule_id",
            "applies_to_question_type",
            "when_question_contains_any",
            "positive_categories",
            "positive_terms",
            "positive_weight",
            "negative_categories",
            "negative_terms",
            "negative_weight",
            "enabled",
        ),
        root_key="rules",
        list_columns=frozenset(
            {
                "when_question_contains_any",
                "positive_categories",
                "positive_terms",
                "negative_categories",
                "negative_terms",
            }
        ),
        int_columns=frozenset({"positive_weight", "negative_weight"}),
    ),
    ExportSpec(
        source_name="06_validation_rules.xlsx",
        output_name="validation_rules.json",
        required_columns=(
            "rule_id",
            "applies_to_question_type",
            "when_question_contains_any",
            "answer_must_contain_any",
            "answer_must_contain_all",
            "answer_should_not_contain_any",
            "fallback_if_failed",
            "enabled",
        ),
        root_key="rules",
        list_columns=frozenset(
            {
                "when_question_contains_any",
                "answer_must_contain_any",
                "answer_must_contain_all",
                "answer_should_not_contain_any",
            }
        ),
        bool_columns=frozenset({"fallback_if_failed", "enabled"}),
    ),
    ExportSpec(
        source_name="07_answer_templates.xlsx",
        output_name="answer_templates.json",
        required_columns=(
            "template_id",
            "applies_to_question_type",
            "format",
            "instruction",
            "max_bullets",
            "enabled",
        ),
        root_key="templates",
        int_columns=frozenset({"max_bullets"}),
    ),
    ExportSpec(
        source_name="08_golden_qa.xlsx",
        output_name="golden_qa.jsonl",
        required_columns=(
            "case_id",
            "question",
            "expected_all",
            "expected_any",
            "forbidden_any",
            "expected_citation",
            "expected_question_type",
            "priority",
            "enabled",
        ),
        root_key=None,
        list_columns=frozenset({"expected_all", "expected_any", "forbidden_any"}),
        int_columns=frozenset({"priority"}),
        bool_columns=frozenset({"expected_citation", "enabled"}),
        jsonl=True,
    ),
)


TRUE_VALUES = {"true", "yes", "1", "有効"}
FALSE_VALUES = {"false", "no", "0", "無効"}


def cell_is_empty(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    return isinstance(value, str) and value.strip() == ""


def normalize_header(value: Any) -> str:
    return "" if value is None else str(value).strip()


def context(spec: ExportSpec, column: str | None = None, row_number: int | None = None) -> str:
    parts = [spec.source_name]
    if row_number is not None:
        parts.append(f"row {row_number}")
    if column is not None:
        parts.append(f"column '{column}'")
    return ", ".join(parts)


def warn(message: str) -> None:
    print(f"WARNING: {message}", file=sys.stderr)


def parse_bool(value: Any, spec: ExportSpec, column: str, row_number: int, *, strict: bool) -> bool | str:
    if cell_is_empty(value):
        return True if column == "enabled" else ""
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        if value == 1:
            return True
        if value == 0:
            return False
    text = str(value).strip().lower()
    if text in TRUE_VALUES:
        return True
    if text in FALSE_VALUES:
        return False

    message = f"{context(spec, column, row_number)} has invalid boolean value: {value!r}"
    if strict:
        raise ExportError(message)
    warn(f"{message}; exported as empty string")
    return ""


def parse_int(value: Any, spec: ExportSpec, column: str, row_number: int, *, strict: bool) -> int | str:
    if cell_is_empty(value):
        return ""
    if isinstance(value, bool):
        message = f"{context(spec, column, row_number)} has boolean where integer is required: {value!r}"
        if strict:
            raise ExportError(message)
        warn(f"{message}; exported as empty string")
        return ""
    if isinstance(value, int):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    text = str(value).strip()
    try:
        return int(text)
    except ValueError:
        message = f"{context(spec, column, row_number)} has invalid integer value: {value!r}"
        if strict:
            raise ExportError(message)
        warn(f"{message}; exported as empty string")
        return ""


def parse_list(value: Any) -> list[Any]:
    if cell_is_empty(value):
        return []
    if isinstance(value, str):
        return [item.strip() for item in value.split("|") if item.strip()]
    return [value]


def parse_scalar(value: Any) -> Any:
    if cell_is_empty(value):
        return ""
    if isinstance(value, str):
        return value.strip()
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def parse_profile_value(value: Any) -> Any:
    if cell_is_empty(value):
        return ""
    if isinstance(value, bool):
        return value
    if isinstance(value, int) and not isinstance(value, bool):
        return value
    if isinstance(value, float) and value.is_integer():
        return int(value)
    if isinstance(value, str):
        text = value.strip()
        lowered = text.lower()
        if lowered in TRUE_VALUES:
            return True
        if lowered in FALSE_VALUES:
            return False
        try:
            return int(text)
        except ValueError:
            return text
    return value


def convert_cell(value: Any, spec: ExportSpec, column: str, row_number: int, *, strict: bool) -> Any:
    if column in spec.list_columns:
        return parse_list(value)
    if column in spec.int_columns:
        return parse_int(value, spec, column, row_number, strict=strict)
    if column in spec.bool_columns:
        return parse_bool(value, spec, column, row_number, strict=strict)
    return parse_scalar(value)


def is_empty_row(row: Iterable[Any]) -> bool:
    return all(cell_is_empty(value) for value in row)


def load_rows(spec: ExportSpec, input_dir: Path, *, strict: bool) -> list[dict[str, Any]]:
    path = input_dir / spec.source_name
    if not path.exists():
        raise ExportError(f"{spec.source_name} not found in {input_dir}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    if "data" not in workbook.sheetnames:
        raise ExportError(f"{spec.source_name} is missing required sheet 'data'")

    worksheet = workbook["data"]
    rows = worksheet.iter_rows(values_only=True)
    try:
        raw_headers = next(rows)
    except StopIteration as exc:
        raise ExportError(f"{spec.source_name} is empty") from exc

    headers = [normalize_header(value) for value in raw_headers]
    missing_columns = [column for column in spec.required_columns if column not in headers]
    if missing_columns:
        message = f"{spec.source_name} missing required columns: {', '.join(missing_columns)}"
        if strict:
            raise ExportError(message)
        warn(f"{message}; missing values will be exported as empty strings/lists")

    exported_rows: list[dict[str, Any]] = []
    for offset, raw_row in enumerate(rows, start=2):
        if is_empty_row(raw_row):
            continue

        row_by_column = dict(zip(headers, raw_row, strict=False))
        enabled = parse_bool(row_by_column.get("enabled"), spec, "enabled", offset, strict=strict)
        if enabled is False:
            continue

        item: dict[str, Any] = {}
        for column in spec.required_columns:
            if column == "description" and spec.source_name == "01_profile.xlsx":
                continue
            item[column] = convert_cell(row_by_column.get(column), spec, column, offset, strict=strict)

        exported_rows.append(item)

    return exported_rows


def export_profile(rows: list[dict[str, Any]], spec: ExportSpec, *, strict: bool) -> dict[str, Any]:
    profile: dict[str, Any] = {}
    for index, row in enumerate(rows, start=2):
        key = row.get("key", "")
        if not key:
            message = f"{context(spec, 'key', index)} is required"
            if strict:
                raise ExportError(message)
            warn(f"{message}; row skipped")
            continue
        profile[str(key)] = parse_profile_value(row.get("value", ""))
    return profile


def write_json(path: Path, payload: Any) -> None:
    path.write_text(
        json.dumps(payload, ensure_ascii=False, indent=2, allow_nan=False) + "\n",
        encoding="utf-8",
    )


def write_jsonl(path: Path, rows: list[dict[str, Any]]) -> None:
    with path.open("w", encoding="utf-8") as file:
        for row in rows:
            file.write(json.dumps(row, ensure_ascii=False, allow_nan=False) + "\n")


def export_all(input_dir: Path, output_dir: Path, *, strict: bool) -> list[tuple[Path, int]]:
    output_dir.mkdir(parents=True, exist_ok=True)
    results: list[tuple[Path, int]] = []

    for spec in EXPORT_SPECS:
        rows = load_rows(spec, input_dir, strict=strict)
        output_path = output_dir / spec.output_name

        if spec.source_name == "01_profile.xlsx":
            profile = export_profile(rows, spec, strict=strict)
            write_json(output_path, profile)
            count = len(profile)
        elif spec.jsonl:
            write_jsonl(output_path, rows)
            count = len(rows)
        else:
            assert spec.root_key is not None
            write_json(output_path, {spec.root_key: rows})
            count = len(rows)

        results.append((output_path, count))

    return results


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Export RAG profile Excel files to JSON/JSONL.")
    parser.add_argument(
        "--input-dir",
        type=Path,
        default=DEFAULT_INPUT_DIR,
        help=f"Input directory for Excel files. Default: {DEFAULT_INPUT_DIR}",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=DEFAULT_OUTPUT_DIR,
        help=f"Output directory for JSON/JSONL files. Default: {DEFAULT_OUTPUT_DIR}",
    )
    parser.add_argument(
        "--strict",
        action="store_true",
        help="Fail on missing required columns or invalid typed values.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    try:
        results = export_all(args.input_dir, args.output_dir, strict=args.strict)
    except ExportError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    print("Exported files:")
    for path, count in results:
        print(f"- {path}: {count}")
    print("Count summary:")
    for path, count in results:
        print(f"{path.name}: {count}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
