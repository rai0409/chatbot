#!/usr/bin/env python3
import html
import json
import re
import shutil
import subprocess
import tempfile
import xml.etree.ElementTree as ET
from collections import Counter
from pathlib import Path
from typing import Dict, List, Tuple, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


PDF_PATH = Path("pdfs/040219e-biscfaq.pdf")
OUT_DIR = Path("artifacts/qa_pair_extraction_040219_bbox")
OUT_JSONL = OUT_DIR / "qa_pair_cases_040219_bbox.jsonl"
OUT_XLSX = OUT_DIR / "qa_pair_cases_040219_bbox_review.xlsx"
OUT_REPORT = OUT_DIR / "qa_pair_extraction_040219_bbox_report.md"
OUT_DEBUG_XML = OUT_DIR / "debug_bbox_040219.html"
OUT_DEBUG_LINES = OUT_DIR / "debug_bbox_lines_040219.json"
OUT_DEBUG_ROWS = OUT_DIR / "debug_bbox_rows_040219.json"


QUESTION_HINTS = [
    "でしょうか",
    "ですか",
    "ますか",
    "できますか",
    "出来ますか",
    "どこにありますか",
    "何ですか",
    "必要ですか",
    "教えて下さい",
    "教えてください",
    "どうすれば",
    "どの様に",
    "どのように",
    "可能でしょうか",
    "利用できますか",
    "表示されます",
    "エラー",
    "どこから",
    "何枚",
    "何度",
    "何を",
    "なぜ",
]

ANSWER_HINTS = [
    "できます",
    "出来ます",
    "できません",
    "出来ません",
    "必要です",
    "不要です",
    "ご確認下さい",
    "ご確認ください",
    "確認下さい",
    "参照下さい",
    "ご利用下さい",
    "お問い合わせ下さい",
    "お問合せ下さい",
    "お願いします",
    "となります",
    "です",
    "ます",
    "下さい",
]


def norm(s: str) -> str:
    s = html.unescape(s)
    s = s.replace("\r\n", "\n").replace("\r", "\n").replace("\u3000", " ")
    s = re.sub(r"[ \t]+", " ", s)
    s = re.sub(r"\n{3,}", "\n\n", s)
    return s.strip()


def compact(s: str) -> str:
    return re.sub(r"\s+", " ", s.replace("\u3000", " ")).strip()


def run_pdftotext_bbox(pdf: Path) -> str:
    if not shutil.which("pdftotext"):
        raise SystemExit("pdftotext がありません: sudo apt install -y poppler-utils")

    with tempfile.TemporaryDirectory() as td:
        out = Path(td) / "bbox.html"
        cmd = ["pdftotext", "-bbox-layout", "-enc", "UTF-8", str(pdf), str(out)]
        subprocess.run(cmd, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        return out.read_text(encoding="utf-8", errors="replace")


def parse_bbox_words(xml_text: str) -> List[Dict]:
    root = ET.fromstring(xml_text)
    pages = []

    for page_index, page in enumerate([e for e in root.iter() if e.tag.endswith("page")], start=1):
        width = float(page.attrib.get("width", "0"))
        height = float(page.attrib.get("height", "0"))

        words = []
        for w in page.iter():
            if not w.tag.endswith("word"):
                continue
            text = compact("".join(w.itertext()))
            if not text:
                continue
            try:
                x_min = float(w.attrib["xMin"])
                y_min = float(w.attrib["yMin"])
                x_max = float(w.attrib["xMax"])
                y_max = float(w.attrib["yMax"])
            except Exception:
                continue

            words.append({
                "text": text,
                "x_min": x_min,
                "x_max": x_max,
                "y_min": y_min,
                "y_max": y_max,
                "x_mid": (x_min + x_max) / 2.0,
                "y_mid": (y_min + y_max) / 2.0,
            })

        pages.append({
            "page": page_index,
            "width": width,
            "height": height,
            "words": words,
        })

    return pages


def group_words_to_lines(words: List[Dict], y_tol: float = 3.0) -> List[Dict]:
    words = sorted(words, key=lambda w: (w["y_mid"], w["x_min"]))
    groups: List[List[Dict]] = []

    for w in words:
        if not groups:
            groups.append([w])
            continue

        last = groups[-1]
        avg_y = sum(x["y_mid"] for x in last) / len(last)
        if abs(w["y_mid"] - avg_y) <= y_tol:
            last.append(w)
        else:
            groups.append([w])

    lines = []
    for group in groups:
        group = sorted(group, key=lambda w: w["x_min"])
        text = compact(" ".join(w["text"] for w in group))
        if not text:
            continue
        lines.append({
            "y": sum(w["y_mid"] for w in group) / len(group),
            "x_min": min(w["x_min"] for w in group),
            "x_max": max(w["x_max"] for w in group),
            "words": group,
            "text": text,
        })

    return lines


def find_header_boundaries(page: Dict, lines: List[Dict]) -> Optional[List[float]]:
    """
    Header is expected to contain: 質問内容 / 質問 / 回答 / 区分.
    Return boundaries for topic/question/answer/type.
    """
    for line in lines[:30]:
        text = line["text"]
        if ("質問内容" in text or "問内容" in text) and "回答" in text:
            words = line["words"]

            label_positions = {}
            for w in words:
                t = w["text"]
                if "質問内容" in t or "問内容" in t:
                    label_positions["topic"] = w["x_mid"]
                elif t == "質問" or t == "問":
                    label_positions["question"] = w["x_mid"]
                elif "回答" in t:
                    label_positions["answer"] = w["x_mid"]
                elif "区分" in t:
                    label_positions["type"] = w["x_mid"]

            if "topic" in label_positions and "answer" in label_positions:
                topic_x = label_positions["topic"]
                question_x = label_positions.get("question")
                answer_x = label_positions["answer"]
                type_x = label_positions.get("type")

                if question_x is None:
                    # topic and answer exist. Estimate question in between.
                    question_x = (topic_x + answer_x) / 2.0

                if type_x is None:
                    type_x = page["width"] - 40

                b1 = (topic_x + question_x) / 2.0
                b2 = (question_x + answer_x) / 2.0
                b3 = (answer_x + type_x) / 2.0
                return [b1, b2, b3]

    return None


def fallback_boundaries(page: Dict) -> List[float]:
    """
    Approximate columns for e-BISC FAQ layout.
    Usually:
    left: 質問内容
    middle: 質問
    right: 回答
    far right: 区分
    """
    w = page["width"]
    return [w * 0.22, w * 0.52, w * 0.86]


def line_to_columns(line: Dict, boundaries: List[float]) -> Dict[str, str]:
    cols = {"topic": [], "question": [], "answer": [], "type": []}
    b1, b2, b3 = boundaries

    for w in line["words"]:
        x = w["x_mid"]
        if x < b1:
            cols["topic"].append(w["text"])
        elif x < b2:
            cols["question"].append(w["text"])
        elif x < b3:
            cols["answer"].append(w["text"])
        else:
            cols["type"].append(w["text"])

    return {k: compact(" ".join(v)) for k, v in cols.items()}


def is_header_or_footer(text: str) -> bool:
    t = compact(text)
    if not t:
        return True
    if "e-BISC" in t and "FAQ" in t:
        return True
    if "質問内容" in t and "回答" in t:
        return True
    if re.fullmatch(r"\d{1,3}", t):
        return True
    return False


def is_question_like(s: str) -> bool:
    s = compact(s)
    if len(s) < 4:
        return False
    return any(h in s for h in QUESTION_HINTS) or s.endswith(("?", "？"))


def is_answer_like(s: str) -> bool:
    s = compact(s)
    if len(s) < 3:
        return False
    return any(h in s for h in ANSWER_HINTS)


def page_to_column_lines(page: Dict) -> List[Dict]:
    lines = group_words_to_lines(page["words"])
    boundaries = find_header_boundaries(page, lines) or fallback_boundaries(page)

    out = []
    for line in lines:
        if is_header_or_footer(line["text"]):
            continue
        cols = line_to_columns(line, boundaries)
        if not any(cols.values()):
            continue
        out.append({
            "page": page["page"],
            "y": line["y"],
            "text": line["text"],
            "cols": cols,
            "boundaries": boundaries,
        })
    return out


def row_has_new_question(cols: Dict[str, str]) -> bool:
    q = cols.get("question", "")
    a = cols.get("answer", "")
    return bool(is_question_like(q) and a)


def build_rows(column_lines: List[Dict]) -> List[Dict]:
    """
    Group visual lines into QA rows.
    A new row starts when question column contains a question-like text and answer column is nonempty.
    Continuation lines are appended to current row.
    """
    rows = []
    cur = None

    def flush():
        nonlocal cur
        if not cur:
            return
        topic = norm(" ".join(cur["topic_parts"]))
        q = norm(" ".join(cur["question_parts"]))
        a = norm(" ".join(cur["answer_parts"]))
        typ = norm(" ".join(cur["type_parts"]))
        if q and a and len(q) >= 4 and len(a) >= 3:
            rows.append({
                "source_page": cur["source_page"],
                "source_topic": topic,
                "question": q,
                "expected_answer": a,
                "source_type": typ,
                "raw_lines": cur["raw_lines"],
            })
        cur = None

    for line in column_lines:
        cols = line["cols"]
        topic = cols.get("topic", "")
        q = cols.get("question", "")
        a = cols.get("answer", "")
        typ = cols.get("type", "")

        new_q = row_has_new_question(cols)

        if new_q:
            flush()
            cur = {
                "source_page": line["page"],
                "topic_parts": [topic] if topic else [],
                "question_parts": [q] if q else [],
                "answer_parts": [a] if a else [],
                "type_parts": [typ] if typ else [],
                "raw_lines": [line],
            }
            continue

        if cur is None:
            # If question and answer are not aligned on this line, skip until a row starts.
            continue

        # Continuation line. Keep columns separated.
        if topic:
            cur["topic_parts"].append(topic)
        if q:
            cur["question_parts"].append(q)
        if a:
            cur["answer_parts"].append(a)
        if typ:
            cur["type_parts"].append(typ)
        cur["raw_lines"].append(line)

    flush()
    return rows


def clean_extracted_row(r: Dict) -> Optional[Dict]:
    q = norm(r["question"])
    a = norm(r["expected_answer"])
    topic = norm(r.get("source_topic", ""))
    typ = norm(r.get("source_type", ""))

    # Remove obvious column spillovers.
    q = re.sub(r"\s+", " ", q).strip()
    a = re.sub(r"\s+", " ", a).strip()

    # Reject if answer appears to contain a full next question.
    # Not always fatal, but for this script we prefer conservative review candidates.
    if len(q) < 4 or len(a) < 3:
        return None
    if q == a:
        return None

    return {
        "source_page": r["source_page"],
        "source_topic": topic,
        "question": q,
        "expected_answer": a,
        "source_type": typ,
    }


def dedupe(rows: List[Dict]) -> List[Dict]:
    out = []
    seen = set()
    for r in rows:
        cr = clean_extracted_row(r)
        if not cr:
            continue
        key = (
            re.sub(r"\s+", "", cr["question"]),
            re.sub(r"\s+", "", cr["expected_answer"])[:160],
        )
        if key in seen:
            continue
        seen.add(key)
        out.append(cr)
    return out


def infer_question_type(q: str, a: str) -> str:
    t = q + " " + a
    if re.search(r"(何枚|何個|何件|何人|何年|何回|何%|何％|いくつ|枚|個|件|16 桁|16桁|3 回|3回)", q):
        return "count_fact"
    if re.search(r"(どこにありますか|どこに|どこから|どこを|どこで)", q):
        return "location"
    if re.search(r"(とは|何ですか|どんな意味)", q):
        return "definition"
    if re.search(r"(手順|方法|どうすれば|設定|登録|入力|クリック|インストール|アンインストール|更新|確認)", t):
        return "procedure"
    if re.search(r"(エラー|表示されます|表示|接続でき|利用でき|失敗|システムエラー|ページが表示)", t):
        return "troubleshooting"
    return "qa_fact"


def extract_terms(answer: str, max_terms: int = 10) -> List[str]:
    out = []
    for x in re.findall(r"\d+(?:,\d{3})*(?:\.\d+)?(?:枚|個|件|人|年|回|%|％|桁)?", answer):
        if x not in out:
            out.append(x)

    chunks = re.findall(r"[一-龥ァ-ヴーA-Za-z0-9][一-龥ァ-ヴーA-Za-z0-9・ー\-/％%]{2,}", answer)
    stop = {
        "こと", "ため", "もの", "場合", "こちら", "あります", "します",
        "ください", "下さい", "いただき", "おります", "お願い"
    }
    for c in chunks:
        c = c.strip("。、，,.（）()[]「」『』:：")
        if not c or c in stop or len(c) > 45:
            continue
        if c not in out:
            out.append(c)
        if len(out) >= max_terms:
            break
    return out[:max_terms]


def expected_min_hits(qtype: str, terms: List[str]) -> int:
    if not terms:
        return 0
    if qtype in {"procedure", "troubleshooting"}:
        return min(3, len(terms))
    return 1


def make_case_rows(rows: List[Dict]) -> List[Dict]:
    out = []
    for i, r in enumerate(rows, start=1):
        q = r["question"]
        a = r["expected_answer"]
        qtype = infer_question_type(q, a)
        any_terms = extract_terms(a)
        all_terms = []
        if qtype == "count_fact":
            all_terms = re.findall(r"\d+(?:,\d{3})*(?:\.\d+)?", a)[:3]

        out.append({
            "case_id": f"biscfaq_bbox_qa_{i:04d}",
            "review_status": "needs_review",
            "source_pdf": str(PDF_PATH),
            "source_page": r["source_page"],
            "source_no": "",
            "source_topic": r.get("source_topic", ""),
            "question": q,
            "expected_answer": a,
            "source_quote": f"Q: {q}\nA: {a}",
            "question_type": qtype,
            "expected_all": all_terms,
            "expected_any": any_terms,
            "expected_min_hits": expected_min_hits(qtype, any_terms),
            "forbidden_any": [],
            "confidence": "medium",
            "notes": f"pattern=bbox_layout;source_type={r.get('source_type','')}",
        })
    return out


def write_jsonl(rows: List[Dict]) -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    with OUT_JSONL.open("w", encoding="utf-8") as f:
        for r in rows:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")


def write_xlsx(rows: List[Dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "qa_pairs_040219_bbox"

    headers = [
        "case_id", "review_status", "source_pdf", "source_page", "source_no",
        "source_topic", "question", "expected_answer", "source_quote",
        "question_type", "expected_all", "expected_any", "expected_min_hits",
        "forbidden_any", "confidence", "notes",
    ]
    ws.append(headers)

    for r in rows:
        ws.append([
            r["case_id"], r["review_status"], r["source_pdf"], r["source_page"],
            r["source_no"], r["source_topic"], r["question"], r["expected_answer"],
            r["source_quote"], r["question_type"],
            json.dumps(r["expected_all"], ensure_ascii=False),
            json.dumps(r["expected_any"], ensure_ascii=False),
            r["expected_min_hits"],
            json.dumps(r["forbidden_any"], ensure_ascii=False),
            r["confidence"], r["notes"],
        ])

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    widths = {
        "A": 22, "B": 16, "C": 42, "D": 12, "E": 10, "F": 34,
        "G": 90, "H": 105, "I": 105, "J": 18, "K": 35, "L": 35,
        "M": 16, "N": 35, "O": 14, "P": 45,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    wb.save(OUT_XLSX)


def write_report(case_rows: List[Dict], column_lines: List[Dict], raw_rows: List[Dict]) -> None:
    by_qtype = Counter(r["question_type"] for r in case_rows)
    by_page = Counter(str(r["source_page"]) for r in case_rows)

    lines = []
    lines.append("# 040219 BISC FAQ BBox QA Extraction Report")
    lines.append("")
    lines.append("## Summary")
    lines.append("")
    lines.append(f"- source_pdf: {PDF_PATH}")
    lines.append(f"- total_qa_pairs: {len(case_rows)}")
    lines.append(f"- visual_column_lines: {len(column_lines)}")
    lines.append(f"- raw_grouped_rows: {len(raw_rows)}")
    lines.append("")
    lines.append("## Counts by Page")
    lines.append("")
    for k, v in by_page.items():
        lines.append(f"- page {k}: {v}")
    lines.append("")
    lines.append("## Question Type Counts")
    lines.append("")
    for k, v in by_qtype.items():
        lines.append(f"- {k}: {v}")
    lines.append("")
    lines.append("## Output Files")
    lines.append("")
    lines.append(f"- {OUT_JSONL}")
    lines.append(f"- {OUT_XLSX}")
    lines.append(f"- {OUT_REPORT}")
    lines.append(f"- {OUT_DEBUG_XML}")
    lines.append(f"- {OUT_DEBUG_LINES}")
    lines.append(f"- {OUT_DEBUG_ROWS}")
    lines.append("")
    lines.append("## Review Rule")
    lines.append("")
    lines.append("- 1行に複数QAが混ざっていないか確認")
    lines.append("- 正しければ review_status を reviewed")
    lines.append("- 修正して使うなら edited")
    lines.append("- 混在・不正確なら rejected")
    OUT_REPORT.write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    if not PDF_PATH.exists():
        raise SystemExit(f"missing pdf: {PDF_PATH}")

    xml_text = run_pdftotext_bbox(PDF_PATH)
    OUT_DEBUG_XML.write_text(xml_text, encoding="utf-8", errors="replace")

    pages = parse_bbox_words(xml_text)

    all_column_lines = []
    for page in pages:
        all_column_lines.extend(page_to_column_lines(page))

    raw_rows = build_rows(all_column_lines)
    cleaned_rows = dedupe(raw_rows)
    case_rows = make_case_rows(cleaned_rows)

    write_jsonl(case_rows)
    write_xlsx(case_rows)
    OUT_DEBUG_LINES.write_text(json.dumps(all_column_lines, ensure_ascii=False, indent=2), encoding="utf-8")
    OUT_DEBUG_ROWS.write_text(json.dumps(raw_rows, ensure_ascii=False, indent=2), encoding="utf-8")
    write_report(case_rows, all_column_lines, raw_rows)

    summary = {
        "source_pdf": str(PDF_PATH),
        "status": "ok" if case_rows else "no_pairs",
        "total_qa_pairs": len(case_rows),
        "visual_column_lines": len(all_column_lines),
        "raw_grouped_rows": len(raw_rows),
        "question_type_counts": dict(Counter(r["question_type"] for r in case_rows)),
        "output_files": [
            str(OUT_JSONL),
            str(OUT_XLSX),
            str(OUT_REPORT),
            str(OUT_DEBUG_XML),
            str(OUT_DEBUG_LINES),
            str(OUT_DEBUG_ROWS),
        ],
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))

    return 0 if case_rows else 1


if __name__ == "__main__":
    raise SystemExit(main())
