#!/usr/bin/env python3
import argparse
import json
import re
from collections import Counter
from pathlib import Path
from typing import Any, Dict, List, Tuple

from pypdf import PdfReader
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def norm(text: str) -> str:
    text = text.replace("\r\n", "\n").replace("\r", "\n").replace("\u3000", " ")
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text.strip()


def line_norm(text: str) -> str:
    return re.sub(r"\s+", " ", text.replace("\u3000", " ")).strip()


def read_pdf_pages(pdf: Path) -> List[Tuple[int, str]]:
    reader = PdfReader(str(pdf))
    out = []
    for i, page in enumerate(reader.pages, start=1):
        try:
            text = page.extract_text() or ""
        except Exception:
            text = ""
        out.append((i, norm(text)))
    return out


def looks_mojibake(text: str) -> bool:
    if not text:
        return True
    weird_chars = sum(text.count(ch) for ch in ["ͯ", "͠", "ϯ", "ʔ", "ɻ", "Θ", "Λ", "Χ", "η"])
    japanese_chars = len(re.findall(r"[ぁ-んァ-ヴー一-龥]", text))
    return weird_chars > 30 and weird_chars > japanese_chars


def infer_question_type(question: str, answer: str) -> str:
    q = question
    a = answer
    joined = q + "\n" + a

    if re.search(r"(何件|何個|何人|何年|何回|何%|何％|いくつ|どのくらい|上限数|目標|施設程度|サンプル)", q):
        return "count_fact"
    if re.search(r"(とは|何ですか|定義|どのようなシステム)", q):
        return "definition"
    if re.search(r"(手順|方法|どうすれば|どのように|入力作業|発送|作成|提出)", q):
        return "procedure"
    if re.search(r"(含まれ|対象|可能|できます|でしょうか|問題がない|必要がある|想定)", q):
        return "qa_fact"
    if re.search(r"(対策|措置|取組|取り組み|運用|実施|調査|集計|レポーティング)", joined):
        return "measure"
    return "other"


def extract_terms(answer: str, max_terms: int = 10) -> List[str]:
    terms: List[str] = []

    for x in re.findall(r"\d+(?:,\d{3})*(?:\.\d+)?(?:千円|施設|サンプル|問|%|％|年|件|人|回)?", answer):
        if x and x not in terms:
            terms.append(x)

    chunks = re.findall(r"[一-龥ァ-ヴーA-Za-z0-9][一-龥ァ-ヴーA-Za-z0-9・ー\-/％%]{2,}", answer)
    stop = {
        "こと", "ため", "もの", "場合", "こちら", "左記", "とおり",
        "あります", "します", "ください", "いただき", "おります",
    }
    for c in chunks:
        c = c.strip("。、，,.（）()[]「」『』:：")
        if not c or c in stop:
            continue
        if len(c) > 40:
            continue
        if c not in terms:
            terms.append(c)
        if len(terms) >= max_terms:
            break

    return terms[:max_terms]


def source_quote(q: str, a: str, limit: int = 800) -> str:
    s = norm(f"Q: {q}\nA: {a}")
    if len(s) > limit:
        return s[:limit].rstrip() + "..."
    return s


def expected_min_hits(question_type: str, terms: List[str]) -> int:
    if not terms:
        return 0
    if question_type in {"procedure", "measure"}:
        return min(3, len(terms))
    return 1


def make_row(
    idx: int,
    source_pdf: Path,
    source_page: int,
    no: str,
    topic: str,
    question: str,
    answer: str,
    confidence: str = "high",
) -> Dict[str, Any]:
    qt = infer_question_type(question, answer)
    expected_any = extract_terms(answer)
    expected_all = []

    if qt == "count_fact":
        expected_all = re.findall(r"\d+(?:,\d{3})*(?:\.\d+)?", answer)[:3]

    return {
        "case_id": f"qa_pdf_{idx:04d}",
        "source_pdf": str(source_pdf),
        "source_page": source_page,
        "source_no": no,
        "source_topic": topic,
        "question": question,
        "expected_answer": answer,
        "source_quote": source_quote(question, answer),
        "question_type": qt,
        "expected_all": expected_all,
        "expected_any": expected_any,
        "expected_min_hits": expected_min_hits(qt, expected_any),
        "forbidden_any": [],
        "confidence": confidence,
        "review_status": "needs_review",
        "notes": "pattern=numbered_table_qa",
    }


def split_numbered_table_blocks(page_text: str) -> List[Tuple[str, List[str]]]:
    """
    58887_95105_misc.pdf のような形式:
    1
    質問項目...
    質問内容...
    回答...
    2
    ...
    を番号ごとのブロックに分ける。
    """
    lines = [line_norm(x) for x in page_text.splitlines()]
    lines = [x for x in lines if x]

    # ヘッダ除外
    filtered = []
    for line in lines:
        if line in {"№ 質問項目 質問内容 回答", "No 質問項目 質問内容 回答"}:
            continue
        if "質問に対する回答" in line:
            continue
        filtered.append(line)

    blocks: List[Tuple[str, List[str]]] = []
    current_no = None
    current_lines: List[str] = []

    for line in filtered:
        if re.fullmatch(r"\d{1,3}", line):
            if current_no is not None:
                blocks.append((current_no, current_lines))
            current_no = line
            current_lines = []
        else:
            if current_no is not None:
                current_lines.append(line)

    if current_no is not None:
        blocks.append((current_no, current_lines))

    return blocks


def is_question_end(line: str) -> bool:
    return bool(re.search(r"(でしょうか。?|ですか。?|ますか。?|ください。?|願いします。?|認識で良いでしょうか。?|認識でお間違いないでしょうか。?|よろしいでしょうか。?)$", line))


def is_answer_start(line: str) -> bool:
    """
    回答らしい開始行を判定。
    58887 PDFでは質問の直後に回答が続くため、典型的な回答文を拾う。
    """
    patterns = [
        r"^今回は",
        r"^フリーアンサー",
        r"^15問程度",
        r"^設問について",
        r"^御社の",
        r"^カスタマーリングス",
        r"^ご提示いただいた",
        r"^含みません",
        r"^既存の",
        r"^入力作業",
        r"^日本語に加え",
        r"^アンケートについて",
        r"^国内のみ",
        r"^想定している",
        r"^観光案内所",
        r"^プレゼント内容",
        r"^年間で",
        r"^対面での",
        r"^設問の内容",
        r"^必須では",
        r"^デジタルアンケート",
        r"^上記の",
    ]
    return any(re.search(p, line) for p in patterns)


def parse_block_to_qa(no: str, block_lines: List[str]) -> Tuple[str, str, str]:
    """
    returns topic, question, answer
    """
    if not block_lines:
        return "", "", ""

    # 質問項目は、質問文が始まるまでの短い行。
    # 質問文は「でしょうか/ですか/ますか/ください/願いします」等で終わるまで。
    # 回答はその後。
    q_start = None

    for i, line in enumerate(block_lines):
        if re.search(r"(でしょうか|ですか|ますか|ご教示|認識|想定|可能|問題|必要|教えて)", line):
            q_start = i
            break

    if q_start is None:
        # それでもだめなら、長い行を質問開始とする
        for i, line in enumerate(block_lines):
            if len(line) >= 20:
                q_start = i
                break

    if q_start is None:
        return " ".join(block_lines[:3]), "", ""

    topic = " ".join(block_lines[:q_start]).strip()

    # answer_startを探す
    ans_start = None
    for i in range(q_start + 1, len(block_lines)):
        prev_joined = " ".join(block_lines[q_start:i])
        if is_question_end(prev_joined) and (is_answer_start(block_lines[i]) or len(block_lines[i]) >= 5):
            ans_start = i
            break

    # fallback: 質問らしい終端を含む行の次
    if ans_start is None:
        for i in range(q_start, len(block_lines)):
            if is_question_end(" ".join(block_lines[q_start:i + 1])):
                ans_start = i + 1
                break

    if ans_start is None or ans_start >= len(block_lines):
        return topic, " ".join(block_lines[q_start:]).strip(), ""

    question = " ".join(block_lines[q_start:ans_start]).strip()
    answer = " ".join(block_lines[ans_start:]).strip()

    return topic, question, answer


def extract_numbered_table_qa(pdf: Path, pages: List[Tuple[int, str]]) -> List[Dict[str, Any]]:
    rows: List[Dict[str, Any]] = []
    idx = 1

    for page_no, text in pages:
        if "質問内容" not in text or "回答" not in text:
            continue

        blocks = split_numbered_table_blocks(text)
        for no, lines in blocks:
            topic, question, answer = parse_block_to_qa(no, lines)
            if not question or not answer:
                continue

            rows.append(
                make_row(
                    idx=idx,
                    source_pdf=pdf,
                    source_page=page_no,
                    no=no,
                    topic=topic,
                    question=question,
                    answer=answer,
                    confidence="high",
                )
            )
            idx += 1

    return rows


def dedupe(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    seen = set()
    out = []
    for r in rows:
        key = (
            r.get("source_pdf", ""),
            re.sub(r"\s+", "", r.get("question", "")),
            re.sub(r"\s+", "", r.get("expected_answer", ""))[:200],
        )
        if key in seen:
            continue
        seen.add(key)
        out.append(r)

    for i, r in enumerate(out, start=1):
        r["case_id"] = f"qa_pdf_{i:04d}"

    return out


def write_jsonl(rows: List[Dict[str, Any]], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        for r in rows:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")


def write_xlsx(rows: List[Dict[str, Any]], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "qa_pairs"

    headers = [
        "case_id",
        "review_status",
        "source_pdf",
        "source_page",
        "source_no",
        "source_topic",
        "question",
        "expected_answer",
        "source_quote",
        "question_type",
        "expected_all",
        "expected_any",
        "expected_min_hits",
        "forbidden_any",
        "confidence",
        "notes",
    ]
    ws.append(headers)

    for r in rows:
        ws.append([
            r.get("case_id", ""),
            r.get("review_status", ""),
            r.get("source_pdf", ""),
            r.get("source_page", ""),
            r.get("source_no", ""),
            r.get("source_topic", ""),
            r.get("question", ""),
            r.get("expected_answer", ""),
            r.get("source_quote", ""),
            r.get("question_type", ""),
            json.dumps(r.get("expected_all", []), ensure_ascii=False),
            json.dumps(r.get("expected_any", []), ensure_ascii=False),
            r.get("expected_min_hits", 0),
            json.dumps(r.get("forbidden_any", []), ensure_ascii=False),
            r.get("confidence", ""),
            r.get("notes", ""),
        ])

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    widths = {
        "A": 16, "B": 16, "C": 42, "D": 12, "E": 10, "F": 30,
        "G": 75, "H": 85, "I": 85, "J": 18, "K": 35, "L": 35,
        "M": 16, "N": 35, "O": 14, "P": 32,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    wb.save(path)


def write_debug_text(pdf: Path, pages: List[Tuple[int, str]], out_dir: Path) -> None:
    debug_dir = out_dir / "debug_text"
    debug_dir.mkdir(parents=True, exist_ok=True)
    out = debug_dir / f"{pdf.name}.txt"
    parts = []
    for page_no, text in pages:
        parts.append(f"\n\n===== PAGE {page_no} =====\n")
        parts.append(text)
    out.write_text("".join(parts), encoding="utf-8")


def write_report(
    rows: List[Dict[str, Any]],
    pages_by_pdf: Dict[str, int],
    mojibake_pdfs: List[str],
    output_dir: Path,
) -> None:
    path = output_dir / "qa_pair_extraction_report.md"

    by_pdf = Counter(r["source_pdf"] for r in rows)
    by_qt = Counter(r["question_type"] for r in rows)
    by_conf = Counter(r["confidence"] for r in rows)

    lines = []
    lines.append("# QA Pair Extraction Report")
    lines.append("")
    lines.append("## Summary")
    lines.append("")
    lines.append(f"- total_qa_pairs: {len(rows)}")
    lines.append("")
    lines.append("## Pages by PDF")
    lines.append("")
    for pdf, pages in pages_by_pdf.items():
        lines.append(f"- {pdf}: {pages}")
    lines.append("")
    lines.append("## Extracted by PDF")
    lines.append("")
    if by_pdf:
        for pdf, count in by_pdf.items():
            lines.append(f"- {pdf}: {count}")
    else:
        lines.append("- none")
    lines.append("")
    lines.append("## Question Type Counts")
    lines.append("")
    if by_qt:
        for k, v in by_qt.items():
            lines.append(f"- {k}: {v}")
    else:
        lines.append("- none")
    lines.append("")
    lines.append("## Confidence Counts")
    lines.append("")
    if by_conf:
        for k, v in by_conf.items():
            lines.append(f"- {k}: {v}")
    else:
        lines.append("- none")
    lines.append("")
    lines.append("## Mojibake / Extraction Warnings")
    lines.append("")
    if mojibake_pdfs:
        for pdf in mojibake_pdfs:
            lines.append(f"- {pdf}: text appears garbled; QA extraction skipped or unreliable.")
    else:
        lines.append("- none")
    lines.append("")
    lines.append("## Output Files")
    lines.append("")
    lines.append(f"- {output_dir / 'qa_pair_cases.jsonl'}")
    lines.append(f"- {output_dir / 'qa_pair_cases_review.xlsx'}")
    lines.append(f"- {output_dir / 'qa_pair_extraction_report.md'}")
    lines.append("")
    lines.append("## Next")
    lines.append("")
    lines.append("1. `qa_pair_cases_review.xlsx` を確認")
    lines.append("2. 採用する行の `review_status` を `reviewed` に変更")
    lines.append("3. 修正して採用する行は `edited` に変更")
    lines.append("4. 不採用は `rejected` に変更")
    path.write_text("\n".join(lines), encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--pdf", action="append", required=True)
    parser.add_argument("--output-dir", default="artifacts/qa_pair_extraction")
    parser.add_argument("--debug", action="store_true")
    parser.add_argument("--max-pairs", type=int, default=0)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    output_dir = Path(args.output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    all_rows: List[Dict[str, Any]] = []
    pages_by_pdf: Dict[str, int] = {}
    mojibake_pdfs: List[str] = []

    for pdf_str in args.pdf:
        pdf = Path(pdf_str)
        if not pdf.exists():
            mojibake_pdfs.append(f"{pdf} missing")
            continue

        pages = read_pdf_pages(pdf)
        pages_by_pdf[str(pdf)] = len(pages)

        if args.debug:
            write_debug_text(pdf, pages, output_dir)

        full_text = "\n".join(t for _, t in pages)
        if looks_mojibake(full_text):
            mojibake_pdfs.append(str(pdf))
            continue

        rows = extract_numbered_table_qa(pdf, pages)
        all_rows.extend(rows)

    all_rows = dedupe(all_rows)

    if args.max_pairs and len(all_rows) > args.max_pairs:
        all_rows = all_rows[:args.max_pairs]
        for i, r in enumerate(all_rows, start=1):
            r["case_id"] = f"qa_pdf_{i:04d}"

    jsonl_path = output_dir / "qa_pair_cases.jsonl"
    xlsx_path = output_dir / "qa_pair_cases_review.xlsx"

    write_jsonl(all_rows, jsonl_path)
    write_xlsx(all_rows, xlsx_path)
    write_report(all_rows, pages_by_pdf, mojibake_pdfs, output_dir)

    summary = {
        "total_qa_pairs": len(all_rows),
        "by_pdf": dict(Counter(r["source_pdf"] for r in all_rows)),
        "question_type_counts": dict(Counter(r["question_type"] for r in all_rows)),
        "confidence_counts": dict(Counter(r["confidence"] for r in all_rows)),
        "mojibake_or_skipped_pdfs": mojibake_pdfs,
        "output_files": [
            str(jsonl_path),
            str(xlsx_path),
            str(output_dir / "qa_pair_extraction_report.md"),
        ],
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))

    return 0 if all_rows else 2


if __name__ == "__main__":
    raise SystemExit(main())
