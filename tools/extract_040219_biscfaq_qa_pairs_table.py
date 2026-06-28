#!/usr/bin/env python3
import argparse
import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any, Dict, List, Optional

import pdfplumber
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from tools.qa_text_normalization import normalize_display_text


DEFAULT_PDF = Path("pdfs/040219e-biscfaq.pdf")
DEFAULT_OUT_DIR = Path("artifacts/qa_pair_extraction_040219_table")


def norm(s: Any) -> str:
    if s is None:
        return ""
    s = str(s)
    s = s.replace("\r\n", "\n").replace("\r", "\n")
    s = s.replace("\u3000", " ")
    s = re.sub(r"[ \t]+", " ", s)
    s = re.sub(r"\n{2,}", "\n", s)
    return s.strip()


def fix_japanese_spacing(s: str) -> str:
    """
    PDFセル内の折り返しで発生した日本語中の不要スペースを除去する。
    例:
    - ボタ ン -> ボタン
    - 新カード に -> 新カードに
    - 教えて下さ い -> 教えて下さい

    ただし、IC カード / PIN 番号 / URL / Internet Explorer のような
    英数字を含む語間スペースは基本的に保持する。
    """
    jp = r"ぁ-んァ-ヴー一-龥々〆〤"
    s = re.sub(rf"([{jp}])\s+([{jp}])", r"\1\2", s)

    # 日本語の開き括弧・閉じ括弧周辺の余計な空白も軽く整える
    s = re.sub(rf"([{jp}])\s+([。、，．！？!?）」』])", r"\1\2", s)
    s = re.sub(rf"([「『（(])\s+([{jp}])", r"\1\2", s)

    return s


def one_line(s: Any) -> str:
    return normalize_display_text(s)

def is_header_row(cells: List[str]) -> bool:
    joined = " ".join(cells)
    return (
        "質問内容" in joined
        and "質問" in joined
        and "回答" in joined
    ) or joined.strip() in {"質問内容 区分 質問 回答", "区分 質問 回答"}


def is_page_title_row(cells: List[str]) -> bool:
    joined = " ".join(cells)
    if "e-BISC" in joined and "FAQ" in joined:
        return True
    if re.fullmatch(r"\d{1,3}", joined.strip()):
        return True
    return False


def looks_like_question(s: str) -> bool:
    s = one_line(s)
    if not s:
        return False
    hints = [
        "ですか",
        "ますか",
        "でしょうか",
        "出来ますか",
        "できますか",
        "どこにありますか",
        "どこから入れますか",
        "教えて下さい",
        "教えてください",
        "何ですか",
        "とは何",
        "どうしたらいいですか",
        "表示されます",
        "表示される",
        "可能ですか",
        "必要がありますか",
        "必要ですか",
        "確認すればいいですか",
        "入力するのですか",
        "届きません",
        "変わりますか",
        "使えますか",
        "利用出来ますか",
        "利用できますか",
        "選択すればいい",
        "確認出来ますか",
        "確認できますか",
        "作成すればいいですか",
        "提出することは可能",
        "閲覧可能でしょうか",
    ]
    return any(h in s for h in hints) or s.endswith(("?", "？", "。"))


def infer_question_type(question: str, answer: str) -> str:
    q = question
    a = answer
    t = q + " " + a

    if re.search(r"(何枚|何個|何件|何人|何年|何回|何度|何%|何％|いくつ|16 桁|16桁|3 回|3回|500MB|9:00|18:00|24時間)", q + " " + a):
        return "count_fact"
    if re.search(r"(どこにありますか|どこから|どこに|どこを|どこで)", q):
        return "location"
    if re.search(r"(とは何ですか|とはなんですか|とはなんでしょうか|とはど|どんな機能|どの様な機能|何ですか)", q):
        return "definition"
    if re.search(r"(エラー|表示されます|表示される|接続出来ません|失敗しました|使用できません|ログイン|システムエラー|ページが表示)", t):
        return "troubleshooting"
    if re.search(r"(手順|方法|登録|設定|入力|クリック|インストール|アンインストール|更新|確認|提出|印刷|変更)", t):
        return "procedure"
    return "qa_fact"


def extract_terms(answer: str, max_terms: int = 10) -> List[str]:
    out: List[str] = []

    for x in re.findall(r"\d+(?:,\d{3})*(?:\.\d+)?(?:枚|個|件|人|年|回|桁|時間|MB|KB|%|％)?", answer):
        if x not in out:
            out.append(x)

    chunks = re.findall(
        r"[一-龥ァ-ヴーA-Za-z0-9][一-龥ァ-ヴーA-Za-z0-9・ー\-/％%\.]{2,}",
        answer,
    )
    stop = {
        "こと", "ため", "もの", "場合", "こちら", "あります", "します",
        "ください", "下さい", "いただき", "おります", "お願い", "ご確認",
        "出来ます", "できます", "出来ません", "できません",
    }

    for c in chunks:
        c = c.strip("。、，,.（）()[]「」『』:：")
        if not c or c in stop or len(c) > 45:
            continue
        if c not in out:
            out.append(c)
        if len(out) >= max_terms:
            break

    return out[:max_terms]


def expected_min_hits(qtype: str, xs: List[str]) -> int:
    if not xs:
        return 0
    if qtype in {"procedure", "troubleshooting"}:
        return min(3, len(xs))
    return 1


def normalize_table_row(raw_row: List[Any]) -> List[str]:
    return [one_line(c) for c in raw_row]


def parse_table_row(
    cells: List[str],
    current_topic: str,
) -> Optional[Dict[str, str]]:
    """
    想定列:
    - 4列: 質問内容 / 区分 / 質問 / 回答
    - 3列: 区分 / 質問 / 回答
    - 2列: 質問 / 回答
    pdfplumberの抽出差を吸収する。
    """
    cells = [c for c in cells]
    while cells and cells[-1] == "":
        cells.pop()

    if not cells:
        return None

    if is_header_row(cells) or is_page_title_row(cells):
        return None

    # 4列以上の場合は、最後2列を question / answer と見る。
    if len(cells) >= 4:
        topic_candidate = cells[0]
        category_candidate = cells[1]
        question = cells[2]
        answer = " ".join(cells[3:]).strip()

        topic = topic_candidate or category_candidate or current_topic

    elif len(cells) == 3:
        topic = cells[0] or current_topic
        question = cells[1]
        answer = cells[2]

    elif len(cells) == 2:
        topic = current_topic
        question = cells[0]
        answer = cells[1]

    else:
        return None

    topic = one_line(topic)
    question = one_line(question)
    answer = one_line(answer)

    if not question or not answer:
        return None

    # 明らかに列ズレしているものを除外しすぎないよう、最低限だけ見る
    if len(question) < 3 or len(answer) < 3:
        return None

    return {
        "topic": topic,
        "question": question,
        "answer": answer,
    }


def extract_tables_from_pdf(pdf_path: Path, debug: bool = False) -> Dict[str, Any]:
    table_settings_candidates = [
        {
            "vertical_strategy": "lines",
            "horizontal_strategy": "lines",
            "snap_tolerance": 3,
            "join_tolerance": 3,
            "intersection_tolerance": 5,
            "text_x_tolerance": 1,
            "text_y_tolerance": 3,
        },
        {
            "vertical_strategy": "lines",
            "horizontal_strategy": "lines",
            "snap_tolerance": 5,
            "join_tolerance": 5,
            "intersection_tolerance": 8,
            "text_x_tolerance": 2,
            "text_y_tolerance": 4,
        },
    ]

    best_rows: List[Dict[str, Any]] = []
    debug_pages: List[Dict[str, Any]] = []

    for setting_idx, settings in enumerate(table_settings_candidates, start=1):
        candidate_rows: List[Dict[str, Any]] = []
        page_debug: List[Dict[str, Any]] = []
        current_topic = ""

        with pdfplumber.open(str(pdf_path)) as pdf:
            for page_no, page in enumerate(pdf.pages, start=1):
                tables = page.extract_tables(table_settings=settings)

                page_debug.append({
                    "page": page_no,
                    "tables_count": len(tables or []),
                    "settings_index": setting_idx,
                    "settings": settings,
                })

                for table_idx, table in enumerate(tables or [], start=1):
                    for raw_row in table:
                        cells = normalize_table_row(raw_row)

                        if is_header_row(cells) or is_page_title_row(cells):
                            continue

                        parsed = parse_table_row(cells, current_topic=current_topic)
                        if not parsed:
                            # 区分だけの行ならtopic更新候補
                            non_empty = [c for c in cells if c]
                            if len(non_empty) == 1 and len(non_empty[0]) <= 30:
                                current_topic = non_empty[0]
                            continue

                        if parsed["topic"]:
                            current_topic = parsed["topic"]

                        # 質問らしくないものも一部あるので、完全除外ではなくconfidenceで扱う
                        candidate_rows.append({
                            "source_page": page_no,
                            "table_idx": table_idx,
                            "source_topic": current_topic,
                            "question": parsed["question"],
                            "expected_answer": parsed["answer"],
                            "raw_cells": cells,
                            "settings_index": setting_idx,
                        })

        # より件数が多い設定を採用
        if len(candidate_rows) > len(best_rows):
            best_rows = candidate_rows
            debug_pages = page_debug

    return {
        "rows": best_rows,
        "debug_pages": debug_pages,
    }


def clean_and_dedupe(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    seen = set()

    for r in rows:
        q = one_line(r["question"])
        a = one_line(r["expected_answer"])
        topic = one_line(r.get("source_topic", ""))

        if not q or not a:
            continue

        # ヘッダ混入除外
        if q in {"質問", "問"} or a == "回答":
            continue
        if "質問内容" in q and "回答" in a:
            continue

        # 1問1答として最低限成立しているもの
        if len(q) < 4 or len(a) < 4:
            continue

        key = (
            re.sub(r"\s+", "", q),
            re.sub(r"\s+", "", a)[:200],
        )
        if key in seen:
            continue
        seen.add(key)

        r2 = dict(r)
        r2["question"] = q
        r2["expected_answer"] = a
        r2["source_topic"] = topic
        out.append(r2)

    return out


def make_case_rows(rows: List[Dict[str, Any]], pdf_path: Path) -> List[Dict[str, Any]]:
    cases: List[Dict[str, Any]] = []

    for i, r in enumerate(rows, start=1):
        q = r["question"]
        a = r["expected_answer"]
        qtype = infer_question_type(q, a)
        any_terms = extract_terms(a)
        all_terms = []

        if qtype == "count_fact":
            all_terms = re.findall(r"\d+(?:,\d{3})*(?:\.\d+)?", a)[:3]

        confidence = "high"
        if not looks_like_question(q):
            confidence = "medium"
        if len(a) > 1200 or len(q) > 600:
            confidence = "low"

        cases.append({
            "case_id": f"biscfaq_table_qa_{i:04d}",
            "review_status": "needs_review",
            "source_pdf": str(pdf_path),
            "source_page": r.get("source_page", ""),
            "source_no": "",
            "source_topic": r.get("source_topic", ""),
            "question": q,
            "expected_answer": a,
            "source_quote": f"Q: {q}\nA: {a}",
            "question_type": qtype,
            "expected_all": all_terms,
            "expected_any": any_terms,
            "expected_min_hits": expected_min_hits(qtype, any_terms),
            "forbidden_any": [],
            "confidence": confidence,
            "notes": f"pattern=pdfplumber_table;table_idx={r.get('table_idx')};settings={r.get('settings_index')}",
        })

    return cases


def write_jsonl(rows: List[Dict[str, Any]], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as f:
        for r in rows:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")


def write_xlsx(rows: List[Dict[str, Any]], path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "qa_pairs_040219_table"

    headers = [
        "case_id",
        "review_status",
        "source_pdf",
        "source_page",
        "source_no",
        "source_topic",
        "question",
        "expected_answer",
        "source_quote",
        "question_type",
        "expected_all",
        "expected_any",
        "expected_min_hits",
        "forbidden_any",
        "confidence",
        "notes",
    ]
    ws.append(headers)

    for r in rows:
        ws.append([
            r.get("case_id", ""),
            r.get("review_status", ""),
            r.get("source_pdf", ""),
            r.get("source_page", ""),
            r.get("source_no", ""),
            r.get("source_topic", ""),
            r.get("question", ""),
            r.get("expected_answer", ""),
            r.get("source_quote", ""),
            r.get("question_type", ""),
            json.dumps(r.get("expected_all", []), ensure_ascii=False),
            json.dumps(r.get("expected_any", []), ensure_ascii=False),
            r.get("expected_min_hits", 0),
            json.dumps(r.get("forbidden_any", []), ensure_ascii=False),
            r.get("confidence", ""),
            r.get("notes", ""),
        ])

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    widths = {
        "A": 24,
        "B": 16,
        "C": 42,
        "D": 12,
        "E": 10,
        "F": 30,
        "G": 85,
        "H": 105,
        "I": 105,
        "J": 18,
        "K": 35,
        "L": 35,
        "M": 16,
        "N": 35,
        "O": 14,
        "P": 44,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    wb.save(path)


def write_report(
    cases: List[Dict[str, Any]],
    raw_rows: List[Dict[str, Any]],
    debug_pages: List[Dict[str, Any]],
    out_dir: Path,
    pdf_path: Path,
) -> None:
    report_path = out_dir / "qa_pair_extraction_040219_table_report.md"
    by_page = Counter(str(r.get("source_page", "")) for r in cases)
    by_type = Counter(r.get("question_type", "unknown") for r in cases)
    by_conf = Counter(r.get("confidence", "unknown") for r in cases)

    lines = []
    lines.append("# 040219 BISC FAQ Table QA Extraction Report")
    lines.append("")
    lines.append("## Summary")
    lines.append("")
    lines.append(f"- source_pdf: {pdf_path}")
    lines.append(f"- raw_rows: {len(raw_rows)}")
    lines.append(f"- total_qa_pairs: {len(cases)}")
    lines.append("")
    lines.append("## Counts by Page")
    lines.append("")
    for k, v in sorted(by_page.items(), key=lambda x: int(x[0]) if x[0].isdigit() else 999):
        lines.append(f"- page {k}: {v}")
    lines.append("")
    lines.append("## Question Type Counts")
    lines.append("")
    for k, v in by_type.items():
        lines.append(f"- {k}: {v}")
    lines.append("")
    lines.append("## Confidence Counts")
    lines.append("")
    for k, v in by_conf.items():
        lines.append(f"- {k}: {v}")
    lines.append("")
    lines.append("## Page Table Debug")
    lines.append("")
    for d in debug_pages:
        lines.append(f"- page {d['page']}: tables={d['tables_count']}, settings={d['settings_index']}")
    lines.append("")
    lines.append("## Output Files")
    lines.append("")
    lines.append(f"- {out_dir / 'qa_pair_cases_040219_table.jsonl'}")
    lines.append(f"- {out_dir / 'qa_pair_cases_040219_table_review.xlsx'}")
    lines.append(f"- {out_dir / 'qa_pair_extraction_040219_table_report.md'}")
    lines.append(f"- {out_dir / 'debug_raw_table_rows_040219.json'}")
    lines.append("")
    lines.append("## Review Rule")
    lines.append("")
    lines.append("- 1行に1つの質問と1つの回答だけが入っているか確認")
    lines.append("- 正しい行は review_status を reviewed")
    lines.append("- 修正して採用する行は edited")
    lines.append("- 混在・不正確な行は rejected")
    report_path.write_text("\n".join(lines), encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--pdf", default=str(DEFAULT_PDF))
    parser.add_argument("--output-dir", default=str(DEFAULT_OUT_DIR))
    parser.add_argument("--debug", action="store_true")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    pdf_path = Path(args.pdf)
    out_dir = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    if not pdf_path.exists():
        raise SystemExit(f"missing pdf: {pdf_path}")

    extracted = extract_tables_from_pdf(pdf_path, debug=args.debug)
    raw_rows = extracted["rows"]
    debug_pages = extracted["debug_pages"]

    cleaned_rows = clean_and_dedupe(raw_rows)
    cases = make_case_rows(cleaned_rows, pdf_path)

    jsonl_path = out_dir / "qa_pair_cases_040219_table.jsonl"
    xlsx_path = out_dir / "qa_pair_cases_040219_table_review.xlsx"
    raw_debug_path = out_dir / "debug_raw_table_rows_040219.json"

    write_jsonl(cases, jsonl_path)
    write_xlsx(cases, xlsx_path)
    raw_debug_path.write_text(json.dumps(raw_rows, ensure_ascii=False, indent=2), encoding="utf-8")
    write_report(cases, raw_rows, debug_pages, out_dir, pdf_path)

    summary = {
        "source_pdf": str(pdf_path),
        "status": "ok" if cases else "no_pairs",
        "raw_rows": len(raw_rows),
        "total_qa_pairs": len(cases),
        "counts_by_page": dict(Counter(str(r.get("source_page", "")) for r in cases)),
        "question_type_counts": dict(Counter(r.get("question_type", "unknown") for r in cases)),
        "confidence_counts": dict(Counter(r.get("confidence", "unknown") for r in cases)),
        "output_files": [
            str(jsonl_path),
            str(xlsx_path),
            str(out_dir / "qa_pair_extraction_040219_table_report.md"),
            str(raw_debug_path),
        ],
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))

    return 0 if cases else 1


if __name__ == "__main__":
    raise SystemExit(main())
