#!/usr/bin/env python3
import json
import re
from pathlib import Path
from collections import Counter

from pypdf import PdfReader
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


PDF_PATH = Path("pdfs/58887_95105_misc.pdf")
OUT_DIR = Path("artifacts/qa_pair_extraction")
OUT_JSONL = OUT_DIR / "qa_pair_cases.jsonl"
OUT_XLSX = OUT_DIR / "qa_pair_cases_review.xlsx"
OUT_REPORT = OUT_DIR / "qa_pair_extraction_report.md"
OUT_DEBUG = OUT_DIR / "debug_58887_blocks.json"


ANSWER_START_PATTERNS = {
    "1": "プレゼンテーション資料は",
    "2": "フリーアンサーも含みます",
    "3": "15問程度の設問には",
    "4": "設問については",
    "5": "御社のシステムを使用した場合",
    "6": "カスタマーリングスのアンケートフォーム",
    "7": "ご提示いただいたような",
    "8": "含みません",
    "9": "既存のシステムを使用するため",
    "10": "入力作業はこちらで",
    "11": "日本語に加え",
    "12": "アンケートについては",
    "13": "国内のみとし",
    "14": "想定している連絡先は",
    "15": "観光案内所や宿泊施設など",
    "16": "プレゼント内容の選定",
    "17": "年間で1万サンプル",
    "18": "対面での打合せについては",
    "19": "設問の内容にもよりますが",
    "20": "必須ではありません",
    "21": "デジタルアンケートを実施することは",
    "22": "上記の類似事業での課題は",
}


def norm(s: str) -> str:
    s = s.replace("\r\n", "\n").replace("\r", "\n").replace("\u3000", " ")
    s = re.sub(r"[ \t]+", " ", s)
    s = re.sub(r"\n{3,}", "\n\n", s)
    return s.strip()


def line_norm(s: str) -> str:
    return re.sub(r"\s+", " ", s.replace("\u3000", " ")).strip()


def read_pdf_text(pdf: Path) -> tuple[int, str]:
    reader = PdfReader(str(pdf))
    pages = []
    for i, page in enumerate(reader.pages, start=1):
        text = page.extract_text() or ""
        pages.append(f"\n\n===== PAGE {i} =====\n{text}")
    return len(reader.pages), norm("\n".join(pages))


def remove_headers(text: str) -> str:
    lines = []
    for raw in text.splitlines():
        line = line_norm(raw)
        if not line:
            continue
        if "観光デジタルアンケート分析業務" in line:
            continue
        if line == "№ 質問項目 質問内容 回答":
            continue
        if line == "No 質問項目 質問内容 回答":
            continue
        if line.startswith("===== PAGE"):
            continue
        lines.append(line)
    return "\n".join(lines)


def normalize_number_boundaries(text: str) -> str:
    """
    pypdfの抽出では、20番の回答の後ろに
    '21 その他 ...' のように同じ行へ混ざることがある。
    1〜22の番号が単独行として扱われるように境界を補正する。
    """
    t = "\n" + text + "\n"

    # 既に単独行の番号はそのまま。
    # 行中に現れる「 21 その他」「 22 その他」を番号行に分離。
    for n in range(1, 23):
        # 番号の前後が数字でない場合だけ分離
        t = re.sub(
            rf"(?<!\d)\s+({n})(?!\d)\s+",
            rf"\n\1\n",
            t,
        )

    # ただし金額 4,972 や 15問などを壊しすぎないため、よくある壊れを戻す
    t = t.replace("4,\n972\n千円", "4,972千円")
    t = t.replace("\n15\n問", "15問")
    t = t.replace("\n20\n施設", "20施設")
    t = t.replace("\n90\n施設", "90施設")
    t = t.replace("\n1\n万サンプル", "1万サンプル")
    return norm(t)


def split_blocks(text: str):
    text = normalize_number_boundaries(text)
    lines = [line_norm(x) for x in text.splitlines()]
    lines = [x for x in lines if x]

    blocks = []
    current_no = None
    current_lines = []

    for line in lines:
        if re.fullmatch(r"\d{1,2}", line):
            n = int(line)
            if 1 <= n <= 22:
                if current_no is not None:
                    blocks.append((current_no, current_lines))
                current_no = str(n)
                current_lines = []
                continue

        if current_no is not None:
            current_lines.append(line)

    if current_no is not None:
        blocks.append((current_no, current_lines))

    return blocks


def split_question_answer(no: str, lines: list[str]):
    joined = " ".join(lines)
    joined = re.sub(r"\s+", " ", joined).strip()

    marker = ANSWER_START_PATTERNS.get(no)
    if not marker:
        return "", "", ""

    pos = joined.find(marker)
    if pos < 0:
        return "", joined, ""

    before = joined[:pos].strip()
    answer = joined[pos:].strip()

    # before = 質問項目 + 質問内容
    # 質問内容の開始を推定
    question_signals = [
        "プレゼンテーション資料について",
        "15問程度",
        "「佐渡市が指定するアンケートシステム」",
        "アンケートシステムについて",
        "貴市より",
        "仕様書内",
        "アンケートの設問",
        "今回のアンケート調査",
        "アンケートの広報物",
        "発送は",
        "市内事業者の連絡先",
        "チラシ等の印刷",
        "プレゼントの佐渡産品",
        "1回あたり",
        "月毎の打合せ",
        "レポーティングの内容",
        "集計方法として",
        "当事業の受託者",
        "同様の施策",
    ]

    q_pos_candidates = [before.find(sig) for sig in question_signals if before.find(sig) >= 0]
    if q_pos_candidates:
        q_pos = min(q_pos_candidates)
        topic = before[:q_pos].strip()
        question = before[q_pos:].strip()
    else:
        # fallback: 長い文が始まる位置
        parts = before.split()
        topic_parts = []
        question_parts = []
        question_started = False
        for p in parts:
            if not question_started and (len(p) >= 18 or any(x in p for x in ["でしょうか", "ですか", "ご教示", "認識"])):
                question_started = True
            if question_started:
                question_parts.append(p)
            else:
                topic_parts.append(p)
        topic = " ".join(topic_parts).strip()
        question = " ".join(question_parts).strip() or before

    question = cleanup_question(question)
    answer = cleanup_answer(answer)
    topic = cleanup_topic(topic)

    return topic, question, answer


def cleanup_topic(s: str) -> str:
    s = re.sub(r"\s+", " ", s).strip()
    return s


def cleanup_question(s: str) -> str:
    s = re.sub(r"\s+", " ", s).strip()
    return s


def cleanup_answer(s: str) -> str:
    s = re.sub(r"\s+", " ", s).strip()

    # 万一、後続番号が混入した場合は切る
    for n in range(1, 23):
        # " 21 その他 ..." のような残骸を切る
        m = re.search(rf"\s{n}\s+(?:その他|仕様書|実施要領)", s)
        if m:
            s = s[:m.start()].strip()
    return s


def infer_question_type(question: str, answer: str) -> str:
    q = question
    a = answer
    t = q + " " + a
    if re.search(r"(何件|何個|何人|何年|何回|何%|何％|いくつ|上限数|目標|施設程度|サンプル|4,972千円|15問|何回程度)", q):
        return "count_fact"
    if re.search(r"(どのようなシステム|とは|何ですか|初めて)", q):
        return "definition"
    if re.search(r"(入力作業|発送|作成|提出|設計|まとめる|形式|選定|購入)", q):
        return "procedure"
    if re.search(r"(可能|含まれ|含まない|対象|問題がない|必要|想定|認識|間違いない|よろしい|含まれます)", q):
        return "qa_fact"
    if re.search(r"(調査|集計|レポーティング|施策|事業|アンケート|広報物)", t):
        return "measure"
    return "other"


def terms(answer: str, max_terms: int = 10):
    out = []
    for x in re.findall(r"\d+(?:,\d{3})*(?:\.\d+)?(?:千円|施設|サンプル|問|%|％|年|件|人|回)?", answer):
        if x not in out:
            out.append(x)

    chunks = re.findall(r"[一-龥ァ-ヴーA-Za-z0-9][一-龥ァ-ヴーA-Za-z0-9・ー\-/％%]{2,}", answer)
    stop = {"こと", "ため", "もの", "場合", "こちら", "左記", "とおり", "あります", "します", "ください", "いただき", "おります"}
    for c in chunks:
        c = c.strip("。、，,.（）()[]「」『』:：")
        if not c or c in stop or len(c) > 40:
            continue
        if c not in out:
            out.append(c)
        if len(out) >= max_terms:
            break
    return out[:max_terms]


def min_hits(qtype: str, xs: list[str]) -> int:
    if not xs:
        return 0
    if qtype in {"procedure", "measure"}:
        return min(3, len(xs))
    return 1


def make_rows(blocks):
    rows = []
    debug = []
    for no, lines in blocks:
        topic, q, a = split_question_answer(no, lines)
        ok = bool(q and a)
        debug.append({"no": no, "ok": ok, "topic": topic, "question": q, "answer": a, "lines": lines})

        if not ok:
            continue

        qtype = infer_question_type(q, a)
        expected_any = terms(a)
        expected_all = []
        if qtype == "count_fact":
            expected_all = re.findall(r"\d+(?:,\d{3})*(?:\.\d+)?", a)[:3]

        rows.append({
            "case_id": f"qa_pdf_{len(rows)+1:04d}",
            "source_pdf": str(PDF_PATH),
            "source_page": guess_page_by_no(no),
            "source_no": no,
            "source_topic": topic,
            "question": q,
            "expected_answer": a,
            "source_quote": f"Q: {q}\nA: {a}",
            "question_type": qtype,
            "expected_all": expected_all,
            "expected_any": expected_any,
            "expected_min_hits": min_hits(qtype, expected_any),
            "forbidden_any": [],
            "confidence": "high",
            "review_status": "needs_review",
            "notes": "pattern=58887_numbered_table_v2",
        })

    return rows, debug


def guess_page_by_no(no: str) -> int:
    n = int(no)
    if 1 <= n <= 6:
        return 1
    if 7 <= n <= 10:
        return 2
    if 11 <= n <= 18:
        return 3
    return 4


def write_jsonl(rows):
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    with OUT_JSONL.open("w", encoding="utf-8") as f:
        for r in rows:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")


def write_xlsx(rows):
    wb = Workbook()
    ws = wb.active
    ws.title = "qa_pairs"

    headers = [
        "case_id", "review_status", "source_pdf", "source_page", "source_no",
        "source_topic", "question", "expected_answer", "source_quote",
        "question_type", "expected_all", "expected_any", "expected_min_hits",
        "forbidden_any", "confidence", "notes",
    ]
    ws.append(headers)

    for r in rows:
        ws.append([
            r["case_id"], r["review_status"], r["source_pdf"], r["source_page"],
            r["source_no"], r["source_topic"], r["question"], r["expected_answer"],
            r["source_quote"], r["question_type"],
            json.dumps(r["expected_all"], ensure_ascii=False),
            json.dumps(r["expected_any"], ensure_ascii=False),
            r["expected_min_hits"],
            json.dumps(r["forbidden_any"], ensure_ascii=False),
            r["confidence"], r["notes"],
        ])

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    widths = {
        "A": 16, "B": 16, "C": 42, "D": 12, "E": 10, "F": 32,
        "G": 85, "H": 95, "I": 95, "J": 18, "K": 35, "L": 35,
        "M": 16, "N": 35, "O": 14, "P": 34,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    wb.save(OUT_XLSX)


def write_report(rows, pages_count, blocks_count, debug):
    by_qtype = Counter(r["question_type"] for r in rows)
    extracted_nos = sorted(int(r["source_no"]) for r in rows)
    missing = [n for n in range(1, 23) if n not in extracted_nos]
    failed = [d for d in debug if not d["ok"]]

    lines = []
    lines.append("# QA Pair Extraction Report")
    lines.append("")
    lines.append("## Summary")
    lines.append("")
    lines.append(f"- source_pdf: {PDF_PATH}")
    lines.append(f"- pages: {pages_count}")
    lines.append(f"- raw_blocks_count: {blocks_count}")
    lines.append(f"- total_qa_pairs: {len(rows)}")
    lines.append(f"- extracted_nos: {extracted_nos}")
    lines.append(f"- missing_nos: {missing}")
    lines.append(f"- failed_blocks: {len(failed)}")
    lines.append("")
    lines.append("## Question Type Counts")
    lines.append("")
    for k, v in by_qtype.items():
        lines.append(f"- {k}: {v}")
    lines.append("")
    lines.append("## Output Files")
    lines.append("")
    lines.append(f"- {OUT_JSONL}")
    lines.append(f"- {OUT_XLSX}")
    lines.append(f"- {OUT_REPORT}")
    lines.append(f"- {OUT_DEBUG}")
    lines.append("")
    lines.append("## Next")
    lines.append("")
    lines.append("1. qa_pair_cases_review.xlsx を確認")
    lines.append("2. 採用する行の review_status を reviewed に変更")
    lines.append("3. 修正して採用する行は edited に変更")
    lines.append("4. 不採用は rejected に変更")
    OUT_REPORT.write_text("\n".join(lines), encoding="utf-8")


def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    pages_count, text = read_pdf_text(PDF_PATH)
    cleaned = remove_headers(text)
    blocks = split_blocks(cleaned)
    rows, debug = make_rows(blocks)

    write_jsonl(rows)
    write_xlsx(rows)
    OUT_DEBUG.write_text(json.dumps(debug, ensure_ascii=False, indent=2), encoding="utf-8")
    write_report(rows, pages_count, len(blocks), debug)

    extracted_nos = sorted(int(r["source_no"]) for r in rows)
    missing = [n for n in range(1, 23) if n not in extracted_nos]

    summary = {
        "source_pdf": str(PDF_PATH),
        "pages": pages_count,
        "raw_blocks_count": len(blocks),
        "total_qa_pairs": len(rows),
        "extracted_nos": extracted_nos,
        "missing_nos": missing,
        "question_type_counts": dict(Counter(r["question_type"] for r in rows)),
        "output_files": [str(OUT_JSONL), str(OUT_XLSX), str(OUT_REPORT), str(OUT_DEBUG)],
    }
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0 if len(rows) == 22 else 1


if __name__ == "__main__":
    raise SystemExit(main())
