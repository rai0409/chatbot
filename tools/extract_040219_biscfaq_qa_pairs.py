#!/usr/bin/env python3
import json
import re
import shutil
import subprocess
import tempfile
from collections import Counter
from pathlib import Path
from typing import Dict, List, Tuple, Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


PDF_PATH = Path("pdfs/040219e-biscfaq.pdf")
OUT_DIR = Path("artifacts/qa_pair_extraction_040219")
OUT_JSONL = OUT_DIR / "qa_pair_cases_040219.jsonl"
OUT_XLSX = OUT_DIR / "qa_pair_cases_040219_review.xlsx"
OUT_REPORT = OUT_DIR / "qa_pair_extraction_040219_report.md"
OUT_DEBUG_TEXT = OUT_DIR / "debug_text_040219.txt"
OUT_DEBUG_BLOCKS = OUT_DIR / "debug_blocks_040219.json"


QUESTION_HINTS = [
    "でしょうか",
    "ですか",
    "ますか",
    "できますか",
    "出来ますか",
    "どこにありますか",
    "何ですか",
    "必要ですか",
    "教えてください",
    "教えて下さい",
    "どうすれば",
    "どのように",
    "可能でしょうか",
    "利用できますか",
    "表示されます",
    "エラー",
]

ANSWER_HINTS = [
    "できます",
    "出来ます",
    "できません",
    "出来ません",
    "必要です",
    "不要です",
    "ご確認ください",
    "確認ください",
    "参照ください",
    "ご利用ください",
    "お問い合わせください",
    "お願いします",
    "となります",
    "です",
    "ます",
]


def norm(s: str) -> str:
    s = s.replace("\r\n", "\n").replace("\r", "\n").replace("\u3000", " ")
    s = re.sub(r"[ \t]+", " ", s)
    s = re.sub(r"\n{3,}", "\n\n", s)
    return s.strip()


def line_norm(s: str) -> str:
    return re.sub(r"\s+", " ", s.replace("\u3000", " ")).strip()


def run_pdftotext(pdf: Path) -> str:
    if not shutil.which("pdftotext"):
        raise SystemExit(
            "pdftotext がありません。先に実行してください: sudo apt install -y poppler-utils"
        )

    with tempfile.TemporaryDirectory() as td:
        out = Path(td) / "out.txt"

        # -layout は表の列位置を保ちやすい
        cmd = ["pdftotext", "-layout", "-enc", "UTF-8", str(pdf), str(out)]
        subprocess.run(cmd, check=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        return out.read_text(encoding="utf-8", errors="replace")


def looks_mojibake(text: str) -> bool:
    if not text.strip():
        return True

    weird = sum(text.count(ch) for ch in ["ͯ", "͠", "ϯ", "ʔ", "ɻ", "Θ", "Λ", "Χ", "η", "�"])
    jp = len(re.findall(r"[ぁ-んァ-ヴー一-龥]", text))
    kana = len(re.findall(r"[ぁ-んァ-ヴー]", text))

    # 正常日本語がほぼ無い、または前回と同じ文字化け記号が多い場合
    if jp < 50:
        return True
    if weird > 20 and weird > jp * 0.2:
        return True
    if kana < 20 and jp < 100:
        return True
    return False


def clean_text(text: str) -> str:
    lines = []
    for raw in text.splitlines():
        line = raw.rstrip()
        compact = line_norm(line)

        if not compact:
            lines.append("")
            continue

        # ページ番号やURL風ヘッダを軽く除外
        if re.fullmatch(r"\d{1,3}", compact):
            continue
        if "FAQ" in compact and len(compact) < 30:
            continue
        if "ＦＡＱ" in compact and len(compact) < 30:
            continue
        if "センターFAQ" in compact and len(compact) < 40:
            continue
        if "問内容" == compact or "問 回答" == compact or "問内容 問 回答" == compact:
            continue

        lines.append(line)

    return "\n".join(lines)


def split_layout_columns(line: str) -> List[str]:
    """
    pdftotext -layout は列間に複数スペースが入りやすい。
    2個以上の空白で列分割する。
    """
    parts = [p.strip() for p in re.split(r"\s{2,}", line.rstrip()) if p.strip()]
    return parts


def is_question_like(s: str) -> bool:
    s = line_norm(s)
    if len(s) < 4:
        return False
    return any(h in s for h in QUESTION_HINTS) or s.endswith(("?", "？"))


def is_answer_like(s: str) -> bool:
    s = line_norm(s)
    if len(s) < 2:
        return False
    return any(h in s for h in ANSWER_HINTS)


def extract_from_three_column_layout(text: str) -> List[Dict]:
    """
    FAQ表が「問内容 / 問 / 回答」の3列に見える場合の抽出。
    各行を複数スペースで列分割し、右2列を question / answer として蓄積する。
    """
    rows = []
    current_topic = ""
    current_q_parts: List[str] = []
    current_a_parts: List[str] = []

    def flush():
        nonlocal current_topic, current_q_parts, current_a_parts
        q = norm(" ".join(current_q_parts))
        a = norm(" ".join(current_a_parts))
        topic = norm(current_topic)
        if q and a and len(q) >= 4 and len(a) >= 3:
            rows.append({
                "source_topic": topic,
                "question": q,
                "expected_answer": a,
                "pattern": "three_column_layout",
            })
        current_topic = ""
        current_q_parts = []
        current_a_parts = []

    for raw in text.splitlines():
        if not raw.strip():
            continue

        parts = split_layout_columns(raw)

        # 3列以上: topic / question / answer
        if len(parts) >= 3:
            topic, q, a = parts[0], parts[1], " ".join(parts[2:])

            # 新しい質問っぽい行なら前を閉じる
            if current_q_parts and is_question_like(q):
                flush()

            if topic:
                current_topic = (current_topic + " " + topic).strip() if current_topic and not current_q_parts else topic
            if q:
                current_q_parts.append(q)
            if a:
                current_a_parts.append(a)

        # 2列: question / answer または topic / question の可能性
        elif len(parts) == 2:
            left, right = parts

            if is_question_like(left) and is_answer_like(right):
                if current_q_parts:
                    flush()
                current_q_parts.append(left)
                current_a_parts.append(right)
            elif current_q_parts and not current_a_parts:
                current_q_parts.append(left)
                current_a_parts.append(right)
            elif current_q_parts and current_a_parts:
                # 続き行
                if is_question_like(left) and is_answer_like(right):
                    flush()
                    current_q_parts.append(left)
                    current_a_parts.append(right)
                else:
                    current_q_parts.append(left)
                    current_a_parts.append(right)
            else:
                current_topic = (current_topic + " " + left).strip()
                current_q_parts.append(right)

        # 1列: 直前の質問または回答の続き
        elif len(parts) == 1:
            part = parts[0]

            if is_question_like(part) and current_q_parts and current_a_parts:
                flush()
                current_q_parts.append(part)
            elif current_q_parts and not current_a_parts:
                current_q_parts.append(part)
            elif current_q_parts and current_a_parts:
                current_a_parts.append(part)
            else:
                # まだQがない場合はtopic候補
                current_topic = (current_topic + " " + part).strip()

    flush()
    return rows


def extract_from_sequential_faq(text: str) -> List[Dict]:
    """
    列が崩れた場合のfallback。
    質問らしい行を見つけ、次の質問までを回答として取る。
    """
    lines = [line_norm(x) for x in text.splitlines()]
    lines = [x for x in lines if x]

    rows = []
    current_q: Optional[str] = None
    current_a: List[str] = []
    topic_buffer: List[str] = []

    def flush():
        nonlocal current_q, current_a, topic_buffer
        if current_q and current_a:
            rows.append({
                "source_topic": norm(" ".join(topic_buffer[-3:])),
                "question": norm(current_q),
                "expected_answer": norm(" ".join(current_a)),
                "pattern": "sequential_faq",
            })
        current_q = None
        current_a = []

    for line in lines:
        if is_question_like(line):
            if current_q and current_a:
                flush()
            current_q = line
            current_a = []
        else:
            if current_q:
                current_a.append(line)
            else:
                topic_buffer.append(line)

    flush()
    return rows


def dedupe(rows: List[Dict]) -> List[Dict]:
    seen = set()
    out = []
    for r in rows:
        q = re.sub(r"\s+", "", r.get("question", ""))
        a = re.sub(r"\s+", "", r.get("expected_answer", ""))[:200]
        key = (q, a)
        if key in seen:
            continue
        seen.add(key)

        # 明らかに変なものは除外
        if len(r.get("question", "")) < 4:
            continue
        if len(r.get("expected_answer", "")) < 3:
            continue
        if r["question"] == r["expected_answer"]:
            continue

        out.append(r)
    return out


def infer_question_type(q: str, a: str) -> str:
    t = q + " " + a
    if re.search(r"(何枚|何個|何件|何人|何年|何回|何%|何％|いくつ|どのくらい|枚|個|件)", q):
        return "count_fact"
    if re.search(r"(どこにありますか|どこに|どこを|どこで)", q):
        return "location"
    if re.search(r"(とは|何ですか|どんな意味)", q):
        return "definition"
    if re.search(r"(手順|方法|どうすれば|設定|登録|入力|クリック|インストール|アンインストール)", t):
        return "procedure"
    if re.search(r"(エラー|表示されます|表示|接続でき|利用でき)", t):
        return "troubleshooting"
    return "qa_fact"


def extract_terms(answer: str, max_terms: int = 10) -> List[str]:
    out = []
    for x in re.findall(r"\d+(?:,\d{3})*(?:\.\d+)?(?:枚|個|件|人|年|回|%|％)?", answer):
        if x not in out:
            out.append(x)

    chunks = re.findall(r"[一-龥ァ-ヴーA-Za-z0-9][一-龥ァ-ヴーA-Za-z0-9・ー\-/％%]{2,}", answer)
    stop = {"こと", "ため", "もの", "場合", "こちら", "あります", "します", "ください", "下さい", "いただき", "おります"}
    for c in chunks:
        c = c.strip("。、，,.（）()[]「」『』:：")
        if not c or c in stop or len(c) > 40:
            continue
        if c not in out:
            out.append(c)
        if len(out) >= max_terms:
            break
    return out[:max_terms]


def make_rows(extracted: List[Dict]) -> List[Dict]:
    rows = []
    for i, r in enumerate(extracted, start=1):
        q = r["question"]
        a = r["expected_answer"]
        qtype = infer_question_type(q, a)
        expected_any = extract_terms(a)
        expected_all = []
        if qtype == "count_fact":
            expected_all = re.findall(r"\d+(?:,\d{3})*(?:\.\d+)?", a)[:3]

        rows.append({
            "case_id": f"biscfaq_qa_{i:04d}",
            "review_status": "needs_review",
            "source_pdf": str(PDF_PATH),
            "source_page": "",
            "source_no": "",
            "source_topic": r.get("source_topic", ""),
            "question": q,
            "expected_answer": a,
            "source_quote": f"Q: {q}\nA: {a}",
            "question_type": qtype,
            "expected_all": expected_all,
            "expected_any": expected_any,
            "expected_min_hits": 1 if expected_any else 0,
            "forbidden_any": [],
            "confidence": "medium",
            "notes": f"pattern={r.get('pattern', '')}",
        })
    return rows


def write_jsonl(rows: List[Dict]) -> None:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    with OUT_JSONL.open("w", encoding="utf-8") as f:
        for r in rows:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")


def write_xlsx(rows: List[Dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "qa_pairs_040219"

    headers = [
        "case_id",
        "review_status",
        "source_pdf",
        "source_page",
        "source_no",
        "source_topic",
        "question",
        "expected_answer",
        "source_quote",
        "question_type",
        "expected_all",
        "expected_any",
        "expected_min_hits",
        "forbidden_any",
        "confidence",
        "notes",
    ]
    ws.append(headers)

    for r in rows:
        ws.append([
            r["case_id"],
            r["review_status"],
            r["source_pdf"],
            r["source_page"],
            r["source_no"],
            r["source_topic"],
            r["question"],
            r["expected_answer"],
            r["source_quote"],
            r["question_type"],
            json.dumps(r["expected_all"], ensure_ascii=False),
            json.dumps(r["expected_any"], ensure_ascii=False),
            r["expected_min_hits"],
            json.dumps(r["forbidden_any"], ensure_ascii=False),
            r["confidence"],
            r["notes"],
        ])

    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9EAF7")
        cell.alignment = Alignment(wrap_text=True, vertical="top")

    widths = {
        "A": 18, "B": 16, "C": 42, "D": 12, "E": 10, "F": 34,
        "G": 85, "H": 95, "I": 95, "J": 18, "K": 35, "L": 35,
        "M": 16, "N": 35, "O": 14, "P": 36,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    ws.freeze_panes = "A2"
    wb.save(OUT_XLSX)


def write_report(rows: List[Dict], status: str, message: str) -> None:
    by_qtype = Counter(r["question_type"] for r in rows)
    by_pattern = Counter(r["notes"] for r in rows)

    lines = []
    lines.append("# 040219 BISC FAQ QA Extraction Report")
    lines.append("")
    lines.append("## Summary")
    lines.append("")
    lines.append(f"- source_pdf: {PDF_PATH}")
    lines.append(f"- status: {status}")
    lines.append(f"- message: {message}")
    lines.append(f"- total_qa_pairs: {len(rows)}")
    lines.append("")
    lines.append("## Question Type Counts")
    lines.append("")
    if by_qtype:
        for k, v in by_qtype.items():
            lines.append(f"- {k}: {v}")
    else:
        lines.append("- none")
    lines.append("")
    lines.append("## Pattern Counts")
    lines.append("")
    if by_pattern:
        for k, v in by_pattern.items():
            lines.append(f"- {k}: {v}")
    else:
        lines.append("- none")
    lines.append("")
    lines.append("## Output Files")
    lines.append("")
    lines.append(f"- {OUT_JSONL}")
    lines.append(f"- {OUT_XLSX}")
    lines.append(f"- {OUT_REPORT}")
    lines.append(f"- {OUT_DEBUG_TEXT}")
    lines.append(f"- {OUT_DEBUG_BLOCKS}")
    lines.append("")
    lines.append("## Next")
    lines.append("")
    lines.append("1. `qa_pair_cases_040219_review.xlsx` を開く")
    lines.append("2. question / expected_answer の対応を確認する")
    lines.append("3. 正しい行は review_status を reviewed にする")
    lines.append("4. 修正して採用する行は edited にする")
    lines.append("5. 不採用は rejected にする")
    OUT_REPORT.write_text("\n".join(lines), encoding="utf-8")


def main() -> int:
    OUT_DIR.mkdir(parents=True, exist_ok=True)

    if not PDF_PATH.exists():
        raise SystemExit(f"missing pdf: {PDF_PATH}")

    raw_text = run_pdftotext(PDF_PATH)
    OUT_DEBUG_TEXT.write_text(raw_text, encoding="utf-8", errors="replace")

    if looks_mojibake(raw_text):
        write_jsonl([])
        write_xlsx([])
        OUT_DEBUG_BLOCKS.write_text("[]", encoding="utf-8")
        write_report(
            [],
            status="failed_mojibake",
            message="pdftotextでも正常な日本語テキストを取得できませんでした。OCRまたは別方式が必要です。",
        )
        print(json.dumps({
            "source_pdf": str(PDF_PATH),
            "status": "failed_mojibake",
            "total_qa_pairs": 0,
            "debug_text": str(OUT_DEBUG_TEXT),
            "report": str(OUT_REPORT),
            "message": "pdftotextでも文字化けしています。debug_text_040219.txt を確認してください。",
        }, ensure_ascii=False, indent=2))
        return 2

    cleaned = clean_text(raw_text)

    extracted = extract_from_three_column_layout(cleaned)
    extracted = dedupe(extracted)

    # three_columnでほとんど取れない場合だけfallback
    if len(extracted) < 5:
        extracted2 = extract_from_sequential_faq(cleaned)
        extracted = dedupe(extracted + extracted2)

    rows = make_rows(extracted)

    write_jsonl(rows)
    write_xlsx(rows)
    OUT_DEBUG_BLOCKS.write_text(json.dumps(extracted, ensure_ascii=False, indent=2), encoding="utf-8")
    write_report(
        rows,
        status="ok" if rows else "no_pairs",
        message="QA候補を抽出しました。" if rows else "正常テキストは取れましたが、QAペア抽出が0件でした。debug_textを確認してください。",
    )

    print(json.dumps({
        "source_pdf": str(PDF_PATH),
        "status": "ok" if rows else "no_pairs",
        "total_qa_pairs": len(rows),
        "question_type_counts": dict(Counter(r["question_type"] for r in rows)),
        "output_files": [
            str(OUT_JSONL),
            str(OUT_XLSX),
            str(OUT_REPORT),
            str(OUT_DEBUG_TEXT),
            str(OUT_DEBUG_BLOCKS),
        ],
    }, ensure_ascii=False, indent=2))

    return 0 if rows else 1


if __name__ == "__main__":
    raise SystemExit(main())
